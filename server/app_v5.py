"""
硬件监控系统服务端 v5.0 - MySQL版
Flask Web应用,提供API和Web管理界面
支持高并发采集(1000+客户端)
使用MySQL数据库支持高并发
"""

import os
import json
import time
from datetime import datetime
from flask import Flask, request, jsonify, render_template, send_file, session, redirect, url_for
from flask_cors import CORS
import io
import csv
import requests
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import functools
import pymysql
from dbutils.pooled_db import PooledDB

app = Flask(__name__)
CORS(app)

# 登录配置
app.secret_key = 'hardware_monitor_secret_key_2026'
LOGIN_CONFIG = {
    'username': 'xapi',
    'password': 'Ai78965'
}

# MySQL数据库配置
MYSQL_CONFIG = {
    'host': '192.168.20.17',
    'port': 3306,
    'user': 'HwMon',
    'password': 'kk7cy7SDWDMXC5XQ',
    'database': 'hwmon',
    'charset': 'utf8mb4',
    'cursorclass': pymysql.cursors.DictCursor
}

# 数据库连接池
db_pool = None

# 并发采集配置
COLLECT_CONFIG = {
    'max_workers': 50,
    'timeout': 15,
    'retry_times': 0,
}


def init_db_pool():
    """初始化数据库连接池"""
    global db_pool
    db_pool = PooledDB(
        creator=pymysql,
        maxconnections=50,
        mincached=10,
        maxcached=20,
        maxusage=1000,
        blocking=True,
        **MYSQL_CONFIG
    )
    print("MySQL连接池初始化成功")


def get_db():
    """从连接池获取数据库连接"""
    return db_pool.connection()


def init_tables():
    """初始化数据库表结构"""
    conn = get_db()
    try:
        cursor = conn.cursor()
        
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS groups (
                id INT AUTO_INCREMENT PRIMARY KEY,
                name VARCHAR(255) NOT NULL UNIQUE,
                description TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        ''')

        cursor.execute('''
            CREATE TABLE IF NOT EXISTS clients (
                id INT AUTO_INCREMENT PRIMARY KEY,
                client_id VARCHAR(255) NOT NULL UNIQUE,
                hostname VARCHAR(255),
                local_ip VARCHAR(45),
                group_id INT,
                last_report DATETIME,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (group_id) REFERENCES groups(id) ON DELETE SET NULL,
                INDEX idx_client_id (client_id),
                INDEX idx_group_id (group_id),
                INDEX idx_last_report (last_report)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        ''')

        cursor.execute('''
            CREATE TABLE IF NOT EXISTS hardware_reports (
                id INT AUTO_INCREMENT PRIMARY KEY,
                client_id VARCHAR(255) NOT NULL,
                report_data LONGTEXT,
                report_type VARCHAR(50) DEFAULT 'scheduled',
                timestamp DATETIME DEFAULT CURRENT_TIMESTAMP,
                INDEX idx_client_id (client_id),
                INDEX idx_timestamp (timestamp)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        ''')

        cursor.execute('''
            CREATE TABLE IF NOT EXISTS hardware_history (
                id INT AUTO_INCREMENT PRIMARY KEY,
                client_id VARCHAR(255) NOT NULL,
                cpu_info TEXT,
                memory_info TEXT,
                disk_info TEXT,
                gpu_info TEXT,
                snapshot LONGTEXT,
                timestamp DATETIME DEFAULT CURRENT_TIMESTAMP,
                INDEX idx_client_id (client_id),
                INDEX idx_timestamp (timestamp)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        ''')

        cursor.execute('INSERT IGNORE INTO groups (name, description) VALUES (%s, %s)',
                       ('默认分组', '未分组的客户端'))

        cursor.execute('''
            CREATE TABLE IF NOT EXISTS client_baselines (
                client_id VARCHAR(255) PRIMARY KEY,
                cpu_snapshot TEXT,
                gpu_snapshot TEXT,
                memory_snapshot TEXT,
                disk_snapshot TEXT,
                baseline_timestamp TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (client_id) REFERENCES clients(client_id) ON DELETE CASCADE
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        ''')

        cursor.execute('''
            CREATE TABLE IF NOT EXISTS alert_records (
                id INT AUTO_INCREMENT PRIMARY KEY,
                client_id VARCHAR(255) NOT NULL,
                alert_type VARCHAR(100) NOT NULL,
                alert_detail LONGTEXT NOT NULL,
                resolved TINYINT DEFAULT 0,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                INDEX idx_client_id (client_id),
                INDEX idx_resolved (resolved),
                INDEX idx_created_at (created_at)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        ''')

        cursor.execute('''
            CREATE TABLE IF NOT EXISTS email_config (
                id INT PRIMARY KEY,
                smtp_host VARCHAR(255) NOT NULL DEFAULT 'smtp.qq.com',
                smtp_port INT NOT NULL DEFAULT 465,
                smtp_user VARCHAR(255) NOT NULL DEFAULT '',
                smtp_password VARCHAR(255) NOT NULL DEFAULT '',
                sender_name VARCHAR(255) DEFAULT '硬件监控系统',
                recipients TEXT NOT NULL DEFAULT '[]',
                enabled TINYINT DEFAULT 0
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        ''')
        cursor.execute('INSERT IGNORE INTO email_config (id) VALUES (1)')

        cursor.execute('''
            CREATE TABLE IF NOT EXISTS alert_settings (
                id INT PRIMARY KEY,
                monitor_cpu TINYINT DEFAULT 1,
                monitor_gpu TINYINT DEFAULT 1,
                monitor_memory TINYINT DEFAULT 1,
                monitor_disk TINYINT DEFAULT 1,
                monitor_network TINYINT DEFAULT 0,
                monitor_motherboard TINYINT DEFAULT 0,
                monitor_bios TINYINT DEFAULT 0,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        ''')
        cursor.execute('INSERT IGNORE INTO alert_settings (id) VALUES (1)')

        conn.commit()
        print("数据库表初始化完成")
    except Exception as e:
        print(f"数据库初始化失败: {e}")
        raise
    finally:
        conn.close()


def login_required(f):
    """登录验证装饰器"""
    @functools.wraps(f)
    def decorated_function(*args, **kwargs):
        if 'logged_in' not in session:
            if request.path.startswith('/api/'):
                return jsonify({'error': '未登录', 'need_login': True}), 401
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function


def compare_hardware(baseline_snapshots, new_hardware, alert_settings=None):
    """对比新硬件数据与基准快照，返回变更列表"""
    changes = []

    if alert_settings is None:
        alert_settings = {
            'monitor_cpu': 1, 'monitor_gpu': 1, 'monitor_memory': 1,
            'monitor_disk': 1, 'monitor_network': 0,
            'monitor_motherboard': 0, 'monitor_bios': 0
        }

    if alert_settings.get('monitor_cpu', 1):
        old_cpu = json.loads(baseline_snapshots.get('cpu', '[]')) if baseline_snapshots.get('cpu') else []
        new_cpu = new_hardware.get('cpu', [])
        if old_cpu and new_cpu:
            old_names = sorted([c.get('name', '') for c in old_cpu])
            new_names = sorted([c.get('name', '') for c in new_cpu])
            if old_names != new_names:
                changes.append({
                    'type': 'cpu', 'label': 'CPU',
                    'old': ', '.join(old_names) if old_names else '未知',
                    'new': ', '.join(new_names) if new_names else '未知'
                })

    if alert_settings.get('monitor_gpu', 1):
        old_gpu = json.loads(baseline_snapshots.get('gpu', '[]')) if baseline_snapshots.get('gpu') else []
        new_gpu = new_hardware.get('gpu', [])
        if old_gpu and new_gpu:
            old_names = sorted([g.get('name', '') for g in old_gpu])
            new_names = sorted([g.get('name', '') for g in new_gpu])
            if old_names != new_names:
                changes.append({
                    'type': 'gpu', 'label': 'GPU',
                    'old': ', '.join(old_names) if old_names else '未知',
                    'new': ', '.join(new_names) if new_names else '未知'
                })

    if alert_settings.get('monitor_memory', 1):
        old_mem = json.loads(baseline_snapshots.get('memory', '{}')) if baseline_snapshots.get('memory') else {}
        new_mem = new_hardware.get('memory', {})
        if old_mem and new_mem:
            old_total = old_mem.get('total_capacity', 0)
            new_total = new_mem.get('total_capacity', 0)
            if not old_total and old_mem.get('modules'):
                old_total = sum(m.get('capacity', 0) for m in old_mem['modules'])
            if not new_total and new_mem.get('modules'):
                new_total = sum(m.get('capacity', 0) for m in new_mem['modules'])
            if old_total and new_total and old_total != new_total:
                changes.append({
                    'type': 'memory', 'label': '内存',
                    'old': f'{old_total / (1024**3):.1f} GB',
                    'new': f'{new_total / (1024**3):.1f} GB'
                })

    if alert_settings.get('monitor_disk', 1):
        old_disk = json.loads(baseline_snapshots.get('disk', '[]')) if baseline_snapshots.get('disk') else []
        new_disk = new_hardware.get('disk', [])
        if old_disk and new_disk:
            old_info = sorted([(d.get('model', ''), d.get('size', 0)) for d in old_disk])
            new_info = sorted([(d.get('model', ''), d.get('size', 0)) for d in new_disk])
            if len(old_disk) != len(new_disk) or old_info != new_info:
                old_str = ', '.join([f"{d.get('model','?')}({d.get('size',0)//(1024**3)}GB)" for d in old_disk])
                new_str = ', '.join([f"{d.get('model','?')}({d.get('size',0)//(1024**3)}GB)" for d in new_disk])
                changes.append({
                    'type': 'disk', 'label': '硬盘',
                    'old': old_str or '未知', 'new': new_str or '未知'
                })

    return changes


def get_email_config(conn):
    """获取邮件配置"""
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM email_config WHERE id = 1')
    row = cursor.fetchone()
    return dict(row) if row else None


def send_alert_email(client_id, hostname, local_ip, changes):
    """发送硬件变更告警邮件"""
    try:
        conn = get_db()
        config = get_email_config(conn)
        conn.close()

        if not config or not config.get('enabled'):
            return False

        recipients = json.loads(config.get('recipients', '[]'))
        if not recipients:
            return False

        import smtplib
        from email.mime.text import MIMEText
        from email.mime.multipart import MIMEMultipart

        smtp_host = config['smtp_host']
        smtp_port = config['smtp_port']
        smtp_user = config['smtp_user']
        smtp_password = config['smtp_password']
        sender_name = config.get('sender_name', '硬件监控系统')

        change_lines = []
        for c in changes:
            change_lines.append(f'<p><strong>{c["label"]}:</strong> {c["old"]} → {c["new"]}</p>')

        html_body = f'''
        <html><body style="font-family:Microsoft YaHei,Arial,sans-serif;">
        <h2 style="color:#e53e3e;">【硬件变更告警】</h2>
        <table style="border-collapse:collapse;">
            <tr><td style="padding:5px 10px;font-weight:bold;">客户端:</td><td style="padding:5px 10px;">{hostname or client_id}</td></tr>
            <tr><td style="padding:5px 10px;font-weight:bold;">IP地址:</td><td style="padding:5px 10px;">{local_ip or '-'}</td></tr>
            <tr><td style="padding:5px 10px;font-weight:bold;">变更时间:</td><td style="padding:5px 10px;">{datetime.now().strftime("%Y-%m-%d %H:%M:%S")}</td></tr>
        </table>
        <h3 style="margin-top:20px;">变更详情:</h3>
        {''.join(change_lines)}
        <hr style="margin:20px 0;">
        <p style="color:#718096;">如需重置基准，请登录硬件监控系统进行操作。</p>
        </body></html>
        '''

        msg = MIMEMultipart('alternative')
        msg['Subject'] = f'【硬件变更告警】客户端 {hostname or client_id} 检测到硬件变更'
        msg['From'] = f'{sender_name} <{smtp_user}>'
        msg['To'] = ', '.join(recipients)
        msg.attach(MIMEText(html_body, 'html', 'utf-8'))

        server = smtplib.SMTP_SSL(smtp_host, smtp_port)
        server.login(smtp_user, smtp_password)
        server.sendmail(smtp_user, recipients, msg.as_string())
        server.quit()
        return True

    except Exception as e:
        import traceback
        print(f'邮件发送失败: {str(e)}')
        print(traceback.format_exc())
        return False


def collect_single_client(client_id, local_ip):
    """采集单个客户端(用于并发执行)"""
    if not local_ip:
        return {'client_id': client_id, 'status': 'unknown_ip', 'message': 'IP地址未知'}

    try:
        response = requests.post(
            f'http://{local_ip}:13301/api/collect',
            json={'trigger': 'server'},
            timeout=COLLECT_CONFIG['timeout']
        )

        if response.status_code == 200:
            return {'client_id': client_id, 'status': 'success', 'message': '采集成功'}
        else:
            return {'client_id': client_id, 'status': 'failed', 'message': f'HTTP {response.status_code}'}

    except requests.exceptions.Timeout:
        return {'client_id': client_id, 'status': 'timeout', 'message': '连接超时'}
    except requests.exceptions.ConnectionError:
        return {'client_id': client_id, 'status': 'offline', 'message': '无法连接,客户端可能离线或防火墙阻止'}
    except Exception as e:
        return {'client_id': client_id, 'status': 'error', 'message': str(e)}


@app.route('/')
def index():
    """主页"""
    if 'logged_in' not in session:
        return redirect(url_for('login'))
    return render_template('index.html')


@app.route('/login')
def login():
    """登录页面"""
    return render_template('login.html')


@app.route('/api/login', methods=['POST'])
def api_login():
    """登录API"""
    try:
        data = request.json
        username = data.get('username', '')
        password = data.get('password', '')

        if username == LOGIN_CONFIG['username'] and password == LOGIN_CONFIG['password']:
            session['logged_in'] = True
            session['username'] = username
            return jsonify({'status': 'success', 'message': '登录成功'})
        else:
            return jsonify({'status': 'error', 'message': '用户名或密码错误'}), 401

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/logout', methods=['POST'])
def api_logout():
    """登出API"""
    session.clear()
    return jsonify({'status': 'success', 'message': '已登出'})


@app.route('/api/check-login', methods=['GET'])
def check_login():
    """检查登录状态"""
    if 'logged_in' in session:
        return jsonify({'status': 'success', 'logged_in': True, 'username': session.get('username')})
    else:
        return jsonify({'status': 'success', 'logged_in': False})


@app.route('/api/report', methods=['POST'])
def receive_report():
    """接收客户端上报的硬件信息（含基准管理和变更检测）"""
    try:
        data = request.json
        client_id = data.get('client_id')
        hostname = data.get('hostname')
        hardware_info = data.get('hardware_info')
        local_ip = data.get('local_ip', '')
        report_type = data.get('report_type', 'scheduled')

        if not client_id:
            return jsonify({'error': '缺少client_id'}), 400

        conn = get_db()
        try:
            cursor = conn.cursor()

            cursor.execute('''
                INSERT INTO clients (client_id, hostname, local_ip, last_report)
                VALUES (%s, %s, %s, %s)
                ON DUPLICATE KEY UPDATE
                    hostname = VALUES(hostname),
                    local_ip = VALUES(local_ip),
                    last_report = VALUES(last_report)
            ''', (client_id, hostname, local_ip, datetime.now().strftime('%Y-%m-%d %H:%M:%S')))

            cursor.execute('''
                INSERT INTO hardware_reports (client_id, report_data, report_type)
                VALUES (%s, %s, %s)
            ''', (client_id, json.dumps(hardware_info, ensure_ascii=False), report_type))

            cpu_info = json.dumps(hardware_info.get('cpu', []), ensure_ascii=False) if hardware_info.get('cpu') else ''
            mem_info = json.dumps(hardware_info.get('memory', {}), ensure_ascii=False) if hardware_info.get('memory') else ''
            disk_info = json.dumps(hardware_info.get('disk', []), ensure_ascii=False) if hardware_info.get('disk') else ''
            gpu_info = json.dumps(hardware_info.get('gpu', []), ensure_ascii=False) if hardware_info.get('gpu') else ''

            cursor.execute('''
                INSERT INTO hardware_history (client_id, cpu_info, memory_info, disk_info, gpu_info, snapshot)
                VALUES (%s, %s, %s, %s, %s, %s)
            ''', (client_id, cpu_info, mem_info, disk_info, gpu_info, json.dumps(hardware_info, ensure_ascii=False)))

            cursor.execute('''
                DELETE FROM hardware_history
                WHERE client_id = %s AND id NOT IN (
                    SELECT id FROM (
                        SELECT id FROM hardware_history
                        WHERE client_id = %s
                        ORDER BY timestamp DESC
                        LIMIT 10
                    ) AS temp
                )
            ''', (client_id, client_id))

            cursor.execute('SELECT * FROM client_baselines WHERE client_id = %s', (client_id,))
            baseline = cursor.fetchone()

            if not baseline:
                cursor.execute('''
                    INSERT INTO client_baselines (client_id, cpu_snapshot, gpu_snapshot, memory_snapshot, disk_snapshot)
                    VALUES (%s, %s, %s, %s, %s)
                ''', (client_id, cpu_info, gpu_info, mem_info, disk_info))
                print(f'[INFO] 客户端 {client_id} 首次上报，已自动创建基准')
            else:
                cursor.execute('SELECT * FROM alert_settings WHERE id = 1')
                alert_settings_row = cursor.fetchone()
                alert_settings = alert_settings_row if alert_settings_row else None

                baseline_snapshots = {
                    'cpu': baseline['cpu_snapshot'],
                    'gpu': baseline['gpu_snapshot'],
                    'memory': baseline['memory_snapshot'],
                    'disk': baseline['disk_snapshot'],
                    'network': json.dumps(hardware_info.get('network', []), ensure_ascii=False) if hardware_info.get('network') else '',
                    'motherboard': json.dumps(hardware_info.get('motherboard', {}), ensure_ascii=False) if hardware_info.get('motherboard') else '',
                    'bios': json.dumps(hardware_info.get('bios', {}), ensure_ascii=False) if hardware_info.get('bios') else ''
                }

                changes = compare_hardware(baseline_snapshots, hardware_info, alert_settings)

                if changes:
                    alert_detail = json.dumps(changes, ensure_ascii=False)
                    cursor.execute('''
                        INSERT INTO alert_records (client_id, alert_type, alert_detail)
                        VALUES (%s, %s, %s)
                    ''', (client_id, 'hardware_change', alert_detail))

                    print(f'[ALERT] 客户端 {client_id} 检测到硬件变更: {len(changes)} 项')

                    email_sent = send_alert_email(client_id, hostname, local_ip, changes)
                    if email_sent:
                        print(f'[INFO] 已向管理员发送告警邮件')
                    else:
                        print(f'[WARN] 告警邮件发送失败（可能未配置或配置错误）')

            conn.commit()
        finally:
            conn.close()

        return jsonify({'status': 'success', 'message': '接收成功'})

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/clients', methods=['GET'])
@login_required
def get_clients():
    """获取所有客户端列表（支持分组过滤和排序）"""
    try:
        group_id = request.args.get('group_id')
        unassigned = request.args.get('unassigned')
        sort_field = request.args.get('sort_field', 'last_report')
        sort_order = request.args.get('sort_order', 'desc')
        
        conn = get_db()
        try:
            cursor = conn.cursor()

            if unassigned == '1':
                query = '''
                    SELECT c.*, g.name as group_name
                    FROM clients c
                    LEFT JOIN groups g ON c.group_id = g.id
                    WHERE c.group_id IS NULL
                '''
                params = []
            elif group_id:
                query = '''
                    SELECT c.*, g.name as group_name
                    FROM clients c
                    LEFT JOIN groups g ON c.group_id = g.id
                    WHERE c.group_id = %s
                '''
                params = [group_id]
            else:
                query = '''
                    SELECT c.*, g.name as group_name
                    FROM clients c
                    LEFT JOIN groups g ON c.group_id = g.id
                '''
                params = []

            allowed_sort_fields = {
                'hostname': 'c.hostname',
                'local_ip': 'c.local_ip',
                'group_name': 'g.name',
                'last_report': 'c.last_report',
                'status': 'c.last_report'
            }
            
            db_sort_field = allowed_sort_fields.get(sort_field, 'c.last_report')
            order = 'DESC' if sort_order == 'desc' else 'ASC'
            query += f' ORDER BY {db_sort_field} {order}'

            cursor.execute(query, params)
            clients = cursor.fetchall()
        finally:
            conn.close()

        return jsonify({'status': 'success', 'data': clients})

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/client/<client_id>', methods=['GET'])
@login_required
def get_client_detail(client_id):
    """获取客户端详细信息"""
    try:
        conn = get_db()
        try:
            cursor = conn.cursor()

            cursor.execute('''
                SELECT c.*, g.name as group_name
                FROM clients c
                LEFT JOIN groups g ON c.group_id = g.id
                WHERE c.client_id = %s
            ''', (client_id,))

            client = cursor.fetchone()
            if not client:
                return jsonify({'error': '客户端不存在'}), 404

            client_info = dict(client)

            cursor.execute('''
                SELECT report_data, report_type, timestamp
                FROM hardware_reports
                WHERE client_id = %s
                ORDER BY timestamp DESC
                LIMIT 1
            ''', (client_id,))

            report = cursor.fetchone()
            if report:
                client_info['latest_hardware'] = json.loads(report['report_data'])
                client_info['last_hardware_update'] = report['timestamp']
                client_info['last_report_type'] = report['report_type']
        finally:
            conn.close()

        return jsonify({'status': 'success', 'data': client_info})

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/collect/<client_id>', methods=['POST'])
@login_required
def collect_from_client(client_id):
    """主动采集单个客户端"""
    try:
        conn = get_db()
        try:
            cursor = conn.cursor()
            cursor.execute('SELECT local_ip FROM clients WHERE client_id = %s', (client_id,))
            client = cursor.fetchone()

            if not client:
                return jsonify({'error': '客户端不存在'}), 404

            local_ip = client['local_ip']
        finally:
            conn.close()

        if not local_ip:
            return jsonify({'error': '客户端IP地址未知,无法采集'}), 400

        result = collect_single_client(client_id, local_ip)

        if result['status'] == 'success':
            return jsonify({
                'status': 'success',
                'message': f'已向客户端 {client_id} 发送采集请求',
                'client_response': result
            })
        else:
            status_code = 500
            if result['status'] == 'timeout':
                status_code = 408
            elif result['status'] in ('offline', 'unknown_ip'):
                status_code = 404

            return jsonify({
                'status': result['status'],
                'message': result['message']
            }), status_code

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/collect/all', methods=['POST'])
@login_required
def collect_all_clients():
    """一键采集: 并发向所有客户端发送采集请求"""
    try:
        params = request.json or {}
        max_workers = params.get('max_workers', COLLECT_CONFIG['max_workers'])
        timeout = params.get('timeout', COLLECT_CONFIG['timeout'])

        old_workers = COLLECT_CONFIG['max_workers']
        old_timeout = COLLECT_CONFIG['timeout']
        COLLECT_CONFIG['max_workers'] = max_workers
        COLLECT_CONFIG['timeout'] = timeout

        conn = get_db()
        try:
            cursor = conn.cursor()
            cursor.execute('SELECT client_id, local_ip FROM clients')
            clients = cursor.fetchall()
        finally:
            conn.close()

        if not clients:
            COLLECT_CONFIG['max_workers'] = old_workers
            COLLECT_CONFIG['timeout'] = old_timeout
            return jsonify({
                'status': 'completed',
                'total': 0, 'success': 0, 'failed': 0,
                'results': [], 'elapsed_seconds': 0
            })

        start_time = time.time()
        results = []

        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            futures = {
                executor.submit(collect_single_client, c['client_id'], c['local_ip']): c['client_id']
                for c in clients
            }

            for future in as_completed(futures):
                try:
                    result = future.result()
                    results.append(result)
                except Exception as e:
                    results.append({
                        'client_id': futures[future],
                        'status': 'error',
                        'message': str(e)
                    })

        elapsed = time.time() - start_time
        success_count = sum(1 for r in results if r['status'] == 'success')
        fail_count = len(results) - success_count

        COLLECT_CONFIG['max_workers'] = old_workers
        COLLECT_CONFIG['timeout'] = old_timeout

        return jsonify({
            'status': 'completed',
            'total': len(clients),
            'success': success_count,
            'failed': fail_count,
            'elapsed_seconds': round(elapsed, 2),
            'concurrency': max_workers,
            'results': results
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/groups', methods=['GET'])
@login_required
def get_groups():
    """获取所有分组"""
    try:
        conn = get_db()
        try:
            cursor = conn.cursor()
            cursor.execute('''
                SELECT g.*, COUNT(c.id) as client_count
                FROM groups g
                LEFT JOIN clients c ON g.id = c.group_id
                GROUP BY g.id
                ORDER BY g.name
            ''')
            groups = cursor.fetchall()
        finally:
            conn.close()

        return jsonify({'status': 'success', 'data': groups})

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/groups', methods=['POST'])
@login_required
def create_group():
    """创建新分组"""
    try:
        data = request.json
        name = data.get('name')
        description = data.get('description', '')

        if not name:
            return jsonify({'error': '分组名称不能为空'}), 400

        conn = get_db()
        try:
            cursor = conn.cursor()
            cursor.execute('INSERT INTO groups (name, description) VALUES (%s, %s)',
                           (name, description))
            conn.commit()
            group_id = cursor.lastrowid
        finally:
            conn.close()

        return jsonify({'status': 'success', 'group_id': group_id})

    except pymysql.IntegrityError:
        return jsonify({'error': '分组名称已存在'}), 400
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/groups/<int:group_id>', methods=['PUT'])
@login_required
def update_group(group_id):
    """更新分组"""
    try:
        data = request.json
        name = data.get('name')
        description = data.get('description')

        conn = get_db()
        try:
            cursor = conn.cursor()
            cursor.execute('UPDATE groups SET name = %s, description = %s WHERE id = %s',
                           (name, description, group_id))
            conn.commit()
        finally:
            conn.close()

        return jsonify({'status': 'success'})

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/groups/<int:group_id>', methods=['DELETE'])
@login_required
def delete_group(group_id):
    """删除分组"""
    try:
        conn = get_db()
        try:
            cursor = conn.cursor()
            cursor.execute('SELECT name FROM groups WHERE id = %s', (group_id,))
            group = cursor.fetchone()
            if group and group['name'] == '默认分组':
                return jsonify({'error': '不能删除默认分组'}), 400

            cursor.execute('DELETE FROM groups WHERE id = %s', (group_id,))
            conn.commit()
        finally:
            conn.close()

        return jsonify({'status': 'success'})

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/clients/<client_id>/group', methods=['PUT'])
@login_required
def assign_client_to_group(client_id):
    """将客户端分配到分组"""
    try:
        data = request.json
        group_id = data.get('group_id')

        conn = get_db()
        try:
            cursor = conn.cursor()
            cursor.execute('UPDATE clients SET group_id = %s WHERE client_id = %s',
                           (group_id, client_id))
            conn.commit()
        finally:
            conn.close()

        return jsonify({'status': 'success'})

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/clients/<client_id>', methods=['DELETE'])
@login_required
def delete_client(client_id):
    """删除客户端"""
    try:
        conn = get_db()
        try:
            cursor = conn.cursor()
            cursor.execute('DELETE FROM clients WHERE client_id = %s', (client_id,))
            conn.commit()
        finally:
            conn.close()

        return jsonify({'status': 'success'})

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/export/csv', methods=['GET'])
@login_required
def export_csv():
    """导出所有客户端信息为CSV"""
    try:
        group_id = request.args.get('group_id')
        unassigned = request.args.get('unassigned')
        
        conn = get_db()
        try:
            cursor = conn.cursor()

            if unassigned == '1':
                cursor.execute('''
                    SELECT c.client_id, c.hostname, c.local_ip, g.name as group_name,
                           c.last_report, c.created_at
                    FROM clients c
                    LEFT JOIN groups g ON c.group_id = g.id
                    WHERE c.group_id IS NULL
                    ORDER BY c.last_report DESC
                ''')
            elif group_id:
                cursor.execute('''
                    SELECT c.client_id, c.hostname, c.local_ip, g.name as group_name,
                           c.last_report, c.created_at
                    FROM clients c
                    LEFT JOIN groups g ON c.group_id = g.id
                    WHERE c.group_id = %s
                    ORDER BY c.last_report DESC
                ''', (group_id,))
            else:
                cursor.execute('''
                    SELECT c.client_id, c.hostname, c.local_ip, g.name as group_name,
                           c.last_report, c.created_at
                    FROM clients c
                    LEFT JOIN groups g ON c.group_id = g.id
                    ORDER BY c.last_report DESC
                ''')

            clients = cursor.fetchall()
        finally:
            conn.close()

        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(['客户端ID(主机名)', 'IP地址', '分组', '最后上报时间', '创建时间'])

        for client in clients:
            writer.writerow([
                client['client_id'],
                client['local_ip'] or '-',
                client['group_name'] or '未分组',
                client['last_report'],
                client['created_at']
            ])

        output.seek(0)
        return send_file(
            io.BytesIO(output.getvalue().encode('utf-8-sig')),
            mimetype='text/csv',
            as_attachment=True,
            download_name=f'hardware_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
        )

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/export/json', methods=['GET'])
@login_required
def export_json():
    """导出所有客户端信息为JSON"""
    try:
        group_id = request.args.get('group_id')
        unassigned = request.args.get('unassigned')
        
        conn = get_db()
        try:
            cursor = conn.cursor()

            if unassigned == '1':
                cursor.execute('''
                    SELECT c.*, g.name as group_name
                    FROM clients c
                    LEFT JOIN groups g ON c.group_id = g.id
                    WHERE c.group_id IS NULL
                    ORDER BY c.last_report DESC
                ''')
            elif group_id:
                cursor.execute('''
                    SELECT c.*, g.name as group_name
                    FROM clients c
                    LEFT JOIN groups g ON c.group_id = g.id
                    WHERE c.group_id = %s
                    ORDER BY c.last_report DESC
                ''', (group_id,))
            else:
                cursor.execute('''
                    SELECT c.*, g.name as group_name
                    FROM clients c
                    LEFT JOIN groups g ON c.group_id = g.id
                    ORDER BY c.last_report DESC
                ''')

            clients = cursor.fetchall()

            for client in clients:
                cursor.execute('''
                    SELECT report_data, report_type, timestamp
                    FROM hardware_reports
                    WHERE client_id = %s
                    ORDER BY timestamp DESC
                    LIMIT 1
                ''', (client['client_id'],))

                report = cursor.fetchone()
                if report:
                    client['latest_hardware'] = json.loads(report['report_data'])
                    client['last_hardware_update'] = report['timestamp']
                    client['last_report_type'] = report['report_type']
        finally:
            conn.close()

        return send_file(
            io.BytesIO(json.dumps(clients, ensure_ascii=False, indent=2).encode('utf-8')),
            mimetype='application/json',
            as_attachment=True,
            download_name=f'hardware_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.json'
        )

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/export/excel', methods=['GET'])
@login_required
def export_excel():
    """导出客户端硬件信息为Excel文件"""
    try:
        group_id = request.args.get('group_id')
        client_ids_param = request.args.get('client_ids')
        unassigned = request.args.get('unassigned')

        conn = get_db()
        try:
            cursor = conn.cursor()

            if client_ids_param:
                client_ids = [cid.strip() for cid in client_ids_param.split(',') if cid.strip()]
                placeholders = ','.join(['%s' for _ in client_ids])
                cursor.execute(f'''
                    SELECT c.client_id, c.hostname, c.local_ip, g.name as group_name,
                           c.last_report, c.created_at
                    FROM clients c
                    LEFT JOIN groups g ON c.group_id = g.id
                    WHERE c.client_id IN ({placeholders})
                    ORDER BY c.last_report DESC
                ''', client_ids)
            elif unassigned == '1':
                cursor.execute('''
                    SELECT c.client_id, c.hostname, c.local_ip, g.name as group_name,
                           c.last_report, c.created_at
                    FROM clients c
                    LEFT JOIN groups g ON c.group_id = g.id
                    WHERE c.group_id IS NULL
                    ORDER BY c.last_report DESC
                ''')
            elif group_id:
                cursor.execute('''
                    SELECT c.client_id, c.hostname, c.local_ip, g.name as group_name,
                           c.last_report, c.created_at
                    FROM clients c
                    LEFT JOIN groups g ON c.group_id = g.id
                    WHERE c.group_id = %s
                    ORDER BY c.last_report DESC
                ''', (group_id,))
            else:
                cursor.execute('''
                    SELECT c.client_id, c.hostname, c.local_ip, g.name as group_name,
                           c.last_report, c.created_at
                    FROM clients c
                    LEFT JOIN groups g ON c.group_id = g.id
                    ORDER BY c.last_report DESC
                ''')

            clients = cursor.fetchall()
        finally:
            conn.close()

        wb = Workbook()
        ws = wb.active
        ws.title = '客户端列表'

        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='667eea', end_color='667eea', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')

        headers = ['主机名', '客户端ID', 'IP地址', '分组', '最后上报时间', '创建时间',
                   'CPU', '内存', '硬盘', '显卡']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        for row_idx, client in enumerate(clients, 2):
            client_dict = dict(client)
            client_id = client_dict['client_id']

            conn = get_db()
            try:
                cursor = conn.cursor()
                cursor.execute('''
                    SELECT report_data FROM hardware_reports
                    WHERE client_id = %s ORDER BY timestamp DESC LIMIT 1
                ''', (client_id,))
                report = cursor.fetchone()
            finally:
                conn.close()

            cpu_str = '-'
            mem_str = '-'
            disk_str = '-'
            gpu_str = '-'

            if report:
                hardware = json.loads(report['report_data'])

                if hardware.get('cpu') and isinstance(hardware['cpu'], list):
                    cpu_names = [c.get('name', '?') for c in hardware['cpu']]
                    cpu_cores = [f"{c.get('cores', '?')}核" for c in hardware['cpu']]
                    cpu_str = ' | '.join([f"{n}({cores})" for n, cores in zip(cpu_names, cpu_cores)])

                if hardware.get('memory'):
                    total = hardware['memory'].get('total_capacity', 0)
                    if total:
                        mem_str = f"{total / (1024**3):.1f} GB"
                    elif hardware['memory'].get('modules'):
                        total = sum(m.get('capacity', 0) for m in hardware['memory']['modules'])
                        mem_str = f"{total / (1024**3):.1f} GB"

                if hardware.get('disk') and isinstance(hardware['disk'], list):
                    disk_models = [d.get('model', '?') for d in hardware['disk']]
                    disk_sizes = [f"{d.get('size', 0) / (1024**3):.0f}GB" for d in hardware['disk']]
                    disk_str = ' | '.join([f"{m}({s})" for m, s in zip(disk_models, disk_sizes)])

                if hardware.get('gpu') and isinstance(hardware['gpu'], list):
                    gpu_names = [g.get('name', '?') for g in hardware['gpu']]
                    gpu_str = ' | '.join(gpu_names)

            ws.cell(row=row_idx, column=1, value=client_dict.get('hostname') or client_id)
            ws.cell(row=row_idx, column=2, value=client_id)
            ws.cell(row=row_idx, column=3, value=client_dict.get('local_ip') or '-')
            ws.cell(row=row_idx, column=4, value=client_dict.get('group_name') or '未分组')
            ws.cell(row=row_idx, column=5, value=client_dict.get('last_report') or '-')
            ws.cell(row=row_idx, column=6, value=client_dict.get('created_at') or '-')
            ws.cell(row=row_idx, column=7, value=cpu_str)
            ws.cell(row=row_idx, column=8, value=mem_str)
            ws.cell(row=row_idx, column=9, value=disk_str)
            ws.cell(row=row_idx, column=10, value=gpu_str)

        col_widths = [15, 20, 16, 15, 22, 22, 40, 12, 40, 30]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[chr(64 + i) if i <= 26 else 'A' + chr(64 + i - 26)].width = width

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'hardware_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/clients/batch-group', methods=['PUT'])
@login_required
def batch_assign_group():
    """批量分配客户端到分组"""
    try:
        data = request.json
        client_ids = data.get('client_ids', [])
        group_id = data.get('group_id')

        if not client_ids:
            return jsonify({'error': '请选择要操作的客户端'}), 400

        conn = get_db()
        try:
            cursor = conn.cursor()
            placeholders = ','.join(['%s' for _ in client_ids])
            cursor.execute(f'''
                UPDATE clients SET group_id = %s
                WHERE client_id IN ({placeholders})
            ''', [group_id] + client_ids)

            affected = cursor.rowcount
            conn.commit()
        finally:
            conn.close()

        return jsonify({
            'status': 'success',
            'affected': affected,
            'message': f'成功将 {affected} 个客户端分配到分组'
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/client/<client_id>/history', methods=['GET'])
@login_required
def get_client_history(client_id):
    """获取客户端硬件采集历史记录（最近10条）"""
    try:
        conn = get_db()
        try:
            cursor = conn.cursor()

            cursor.execute('SELECT client_id FROM clients WHERE client_id = %s', (client_id,))
            if not cursor.fetchone():
                return jsonify({'error': '客户端不存在'}), 404

            cursor.execute('''
                SELECT cpu_info, memory_info, disk_info, gpu_info, snapshot, timestamp
                FROM hardware_history
                WHERE client_id = %s
                ORDER BY timestamp DESC
                LIMIT 10
            ''', (client_id,))

            history = []
            for row in cursor.fetchall():
                row_dict = dict(row)
                if row_dict.get('snapshot'):
                    snapshot = json.loads(row_dict['snapshot'])
                else:
                    snapshot = {}
                row_dict['snapshot'] = snapshot
                history.append(row_dict)
        finally:
            conn.close()

        return jsonify({'status': 'success', 'data': history})

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/clients/batch-delete', methods=['DELETE'])
@login_required
def batch_delete_clients():
    """批量删除客户端"""
    try:
        data = request.json
        client_ids = data.get('client_ids', [])

        if not client_ids:
            return jsonify({'error': '请选择要删除的客户端'}), 400

        conn = get_db()
        try:
            cursor = conn.cursor()
            placeholders = ','.join(['%s' for _ in client_ids])
            cursor.execute(f'DELETE FROM clients WHERE client_id IN ({placeholders})', client_ids)
            affected = cursor.rowcount
            conn.commit()
        finally:
            conn.close()

        return jsonify({
            'status': 'success',
            'affected': affected,
            'message': f'成功删除 {affected} 个客户端'
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/client/<client_id>/baseline', methods=['GET'])
@login_required
def get_client_baseline(client_id):
    """获取客户端的硬件基准信息"""
    try:
        conn = get_db()
        try:
            cursor = conn.cursor()
            cursor.execute('SELECT * FROM client_baselines WHERE client_id = %s', (client_id,))
            baseline = cursor.fetchone()

            if not baseline:
                return jsonify({'status': 'not_found', 'message': '该客户端尚未建立基准'})

            baseline_dict = dict(baseline)
            for key in ['cpu_snapshot', 'gpu_snapshot', 'memory_snapshot', 'disk_snapshot']:
                if baseline_dict.get(key):
                    baseline_dict[key] = json.loads(baseline_dict[key])
        finally:
            conn.close()

        return jsonify({'status': 'success', 'data': baseline_dict})

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/client/<client_id>/baseline', methods=['POST'])
@login_required
def set_client_baseline(client_id):
    """手动设置/重置客户端的硬件基准"""
    try:
        conn = get_db()
        try:
            cursor = conn.cursor()

            cursor.execute('SELECT client_id FROM clients WHERE client_id = %s', (client_id,))
            if not cursor.fetchone():
                return jsonify({'error': '客户端不存在'}), 404

            cursor.execute('''
                SELECT report_data FROM hardware_reports
                WHERE client_id = %s ORDER BY timestamp DESC LIMIT 1
            ''', (client_id,))
            report = cursor.fetchone()

            if not report:
                return jsonify({'error': '该客户端尚未上报任何硬件数据'}), 400

            hardware_info = json.loads(report['report_data'])

            cpu_info = json.dumps(hardware_info.get('cpu', []), ensure_ascii=False) if hardware_info.get('cpu') else ''
            mem_info = json.dumps(hardware_info.get('memory', {}), ensure_ascii=False) if hardware_info.get('memory') else ''
            disk_info = json.dumps(hardware_info.get('disk', []), ensure_ascii=False) if hardware_info.get('disk') else ''
            gpu_info = json.dumps(hardware_info.get('gpu', []), ensure_ascii=False) if hardware_info.get('gpu') else ''

            cursor.execute('''
                INSERT INTO client_baselines (client_id, cpu_snapshot, gpu_snapshot, memory_snapshot, disk_snapshot)
                VALUES (%s, %s, %s, %s, %s)
                ON DUPLICATE KEY UPDATE
                    cpu_snapshot = VALUES(cpu_snapshot),
                    gpu_snapshot = VALUES(gpu_snapshot),
                    memory_snapshot = VALUES(memory_snapshot),
                    disk_snapshot = VALUES(disk_snapshot),
                    baseline_timestamp = CURRENT_TIMESTAMP
            ''', (client_id, cpu_info, gpu_info, mem_info, disk_info))

            conn.commit()
        finally:
            conn.close()

        return jsonify({'status': 'success', 'message': '基准设置成功'})

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/client/<client_id>/alerts', methods=['GET'])
@login_required
def get_client_alerts(client_id):
    """获取指定客户端的告警记录"""
    try:
        resolved = request.args.get('resolved')

        conn = get_db()
        try:
            cursor = conn.cursor()

            query = '''
                SELECT a.*, c.hostname, c.local_ip
                FROM alert_records a
                LEFT JOIN clients c ON a.client_id = c.client_id
                WHERE a.client_id = %s
            '''
            params = [client_id]

            if resolved is not None:
                query += ' AND a.resolved = %s'
                params.append(1 if resolved == 'true' else 0)

            query += ' ORDER BY a.created_at DESC'

            cursor.execute(query, params)
            alerts = cursor.fetchall()

            for alert in alerts:
                if alert.get('alert_detail'):
                    alert['alert_detail'] = json.loads(alert['alert_detail'])
        finally:
            conn.close()

        return jsonify({'status': 'success', 'data': alerts})

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/alerts', methods=['GET'])
@login_required
def get_all_alerts():
    """获取所有告警记录（支持分页和过滤）"""
    try:
        page = int(request.args.get('page', 1))
        per_page = int(request.args.get('per_page', 20))
        resolved = request.args.get('resolved')

        conn = get_db()
        try:
            cursor = conn.cursor()

            query = '''
                SELECT a.*, c.hostname, c.local_ip
                FROM alert_records a
                LEFT JOIN clients c ON a.client_id = c.client_id
                WHERE 1=1
            '''
            params = []

            if resolved is not None:
                query += ' AND a.resolved = %s'
                params.append(1 if resolved == 'true' else 0)

            count_query = query.replace('SELECT a.*, c.hostname, c.local_ip', 'SELECT COUNT(*)')
            cursor.execute(count_query, params)
            total = cursor.fetchone()['COUNT(*)']

            query += ' ORDER BY a.created_at DESC LIMIT %s OFFSET %s'
            params.extend([per_page, (page - 1) * per_page])

            cursor.execute(query, params)
            alerts = cursor.fetchall()

            for alert in alerts:
                if alert.get('alert_detail'):
                    alert['alert_detail'] = json.loads(alert['alert_detail'])
        finally:
            conn.close()

        return jsonify({
            'status': 'success',
            'data': alerts,
            'pagination': {
                'page': page,
                'per_page': per_page,
                'total': total,
                'pages': (total + per_page - 1) // per_page
            }
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/alerts/<int:alert_id>', methods=['PUT'])
@login_required
def resolve_alert(alert_id):
    """标记单个告警为已解决"""
    try:
        conn = get_db()
        try:
            cursor = conn.cursor()
            cursor.execute('UPDATE alert_records SET resolved = 1 WHERE id = %s', (alert_id,))

            if cursor.rowcount == 0:
                return jsonify({'error': '告警记录不存在'}), 404

            conn.commit()
        finally:
            conn.close()

        return jsonify({'status': 'success', 'message': '告警已标记为已解决'})

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/alerts/batch-resolve', methods=['PUT'])
@login_required
def batch_resolve_alerts():
    """批量标记告警为已解决"""
    try:
        data = request.json
        alert_ids = data.get('alert_ids', [])

        if not alert_ids:
            return jsonify({'error': '请选择要解决的告警'}), 400

        conn = get_db()
        try:
            cursor = conn.cursor()
            placeholders = ','.join(['%s' for _ in alert_ids])
            cursor.execute(f'UPDATE alert_records SET resolved = 1 WHERE id IN ({placeholders})', alert_ids)
            affected = cursor.rowcount

            conn.commit()
        finally:
            conn.close()

        return jsonify({
            'status': 'success',
            'affected': affected,
            'message': f'成功标记 {affected} 个告警为已解决'
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/email-config', methods=['GET'])
@login_required
def get_email_config_api():
    """获取邮件配置"""
    try:
        conn = get_db()
        try:
            config = get_email_config(conn)
        finally:
            conn.close()

        if not config:
            return jsonify({'error': '邮件配置不存在'}), 404

        config_copy = dict(config)
        if config_copy.get('smtp_password'):
            config_copy['smtp_password'] = '******'

        return jsonify({'status': 'success', 'data': config_copy})

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/email-config', methods=['PUT'])
@login_required
def update_email_config():
    """更新邮件配置"""
    try:
        data = request.json

        conn = get_db()
        try:
            cursor = conn.cursor()

            if data.get('smtp_password') == '******':
                cursor.execute('SELECT smtp_password FROM email_config WHERE id = 1')
                row = cursor.fetchone()
                old_password = row['smtp_password'] if row else ''
                data['smtp_password'] = old_password

            cursor.execute('''
                UPDATE email_config SET
                    smtp_host = %s,
                    smtp_port = %s,
                    smtp_user = %s,
                    smtp_password = %s,
                    sender_name = %s,
                    recipients = %s,
                    enabled = %s
                WHERE id = 1
            ''', (
                data.get('smtp_host', 'smtp.qq.com'),
                data.get('smtp_port', 465),
                data.get('smtp_user', ''),
                data.get('smtp_password', ''),
                data.get('sender_name', '硬件监控系统'),
                json.dumps(data.get('recipients', [])),
                1 if data.get('enabled') else 0
            ))

            conn.commit()
        finally:
            conn.close()

        return jsonify({'status': 'success', 'message': '邮件配置已更新'})

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/email-config/test', methods=['POST'])
@login_required
def test_email_config():
    """测试邮件配置"""
    try:
        data = request.json

        import smtplib
        from email.mime.text import MIMEText
        from email.mime.multipart import MIMEMultipart

        smtp_host = data.get('smtp_host', 'smtp.qq.com')
        smtp_port = data.get('smtp_port', 465)
        smtp_user = data.get('smtp_user', '')
        smtp_password = data.get('smtp_password', '')
        sender_name = data.get('sender_name', '硬件监控系统')
        test_recipient = data.get('test_recipient', '')

        if not smtp_user or not smtp_password or not test_recipient:
            return jsonify({'error': '请填写完整的SMTP配置和测试收件人'}), 400

        html_body = f'''
        <html><body style="font-family:Microsoft YaHei,Arial,sans-serif;">
        <h2 style="color:#38a169;">【邮件配置测试】</h2>
        <p>这是一封来自硬件监控系统的测试邮件。</p>
        <p>如果您收到此邮件，说明SMTP配置正确。</p>
        <p style="color:#718096; margin-top:20px;">发送时间: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}</p>
        </body></html>
        '''

        msg = MIMEMultipart('alternative')
        msg['Subject'] = '【测试邮件】硬件监控系统配置测试'
        msg['From'] = f'{sender_name} <{smtp_user}>'
        msg['To'] = test_recipient
        msg.attach(MIMEText(html_body, 'html', 'utf-8'))

        server = smtplib.SMTP_SSL(smtp_host, smtp_port)
        server.login(smtp_user, smtp_password)
        server.sendmail(smtp_user, [test_recipient], msg.as_string())
        server.quit()

        return jsonify({'status': 'success', 'message': '测试邮件发送成功'})

    except smtplib.SMTPAuthenticationError:
        return jsonify({'error': 'SMTP认证失败，请检查用户名和密码'}), 400
    except smtplib.SMTPConnectError:
        return jsonify({'error': '无法连接到SMTP服务器，请检查主机和端口'}), 400
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'邮件发送失败: {str(e)}'}), 500


@app.route('/api/alert-settings', methods=['GET'])
@login_required
def get_alert_settings():
    """获取告警设置"""
    try:
        conn = get_db()
        try:
            cursor = conn.cursor()
            cursor.execute('SELECT * FROM alert_settings WHERE id = 1')
            row = cursor.fetchone()
        finally:
            conn.close()

        if row:
            return jsonify({'status': 'success', 'data': dict(row)})
        else:
            return jsonify({'status': 'success', 'data': {
                'monitor_cpu': 1, 'monitor_gpu': 1, 'monitor_memory': 1,
                'monitor_disk': 1, 'monitor_network': 0,
                'monitor_motherboard': 0, 'monitor_bios': 0
            }})

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/alert-settings', methods=['PUT'])
@login_required
def update_alert_settings():
    """更新告警设置"""
    try:
        data = request.json

        conn = get_db()
        try:
            cursor = conn.cursor()
            cursor.execute('''
                UPDATE alert_settings SET
                    monitor_cpu = %s,
                    monitor_gpu = %s,
                    monitor_memory = %s,
                    monitor_disk = %s,
                    monitor_network = %s,
                    monitor_motherboard = %s,
                    monitor_bios = %s,
                    updated_at = CURRENT_TIMESTAMP
                WHERE id = 1
            ''', (
                1 if data.get('monitor_cpu') else 0,
                1 if data.get('monitor_gpu') else 0,
                1 if data.get('monitor_memory') else 0,
                1 if data.get('monitor_disk') else 0,
                1 if data.get('monitor_network') else 0,
                1 if data.get('monitor_motherboard') else 0,
                1 if data.get('monitor_bios') else 0
            ))
            conn.commit()
        finally:
            conn.close()

        return jsonify({'status': 'success', 'message': '告警设置已更新'})

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/config', methods=['GET'])
@login_required
def get_collect_config():
    """获取采集配置"""
    return jsonify({
        'status': 'success',
        'data': {
            'max_workers': COLLECT_CONFIG['max_workers'],
            'timeout': COLLECT_CONFIG['timeout'],
            'retry_times': COLLECT_CONFIG['retry_times']
        }
    })


if __name__ == '__main__':
    print("=" * 60)
    print("硬件监控系统服务端 v5.0 - MySQL版")
    print("=" * 60)
    
    print("\n[1/3] 初始化MySQL连接池...")
    init_db_pool()
    
    print("[2/3] 初始化数据库表...")
    init_tables()
    
    print("[3/3] 启动服务...")
    print("=" * 60)
    print("访问地址: http://localhost:5000")
    print(f"并发配置: {COLLECT_CONFIG['max_workers']} workers, {COLLECT_CONFIG['timeout']}s timeout")
    print("=" * 60)

    try:
        from waitress import serve
        print("使用 Waitress WSGI 服务器(生产模式)")
        serve(app, host='0.0.0.0', port=5000, threads=20, connection_limit=2000)
    except ImportError:
        print("Waitress未安装,使用Flask内置服务器(开发模式)")
        print("提示: pip install waitress 以启用高并发支持")
        app.run(host='0.0.0.0', port=5000, debug=True)
