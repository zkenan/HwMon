"""
硬件监控系统服务端 - HwMonServer
Flask Web应用,提供API和Web管理界面
支持高并发采集(1000+客户端)
使用MySQL数据库支持高并发
"""

import os
import sys
import json
import time
import logging
from datetime import datetime, timezone, timedelta

# 东八区固定时区偏移（确保所有时间戳使用 UTC+8，避免系统时区不一致问题）
TZ_CST = timezone(timedelta(hours=8), name='CST')
from flask import Flask, request, jsonify, render_template, send_file, session, redirect, url_for
from flask_cors import CORS
import io
import csv
import requests
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import functools
import pymysql
from dbutils.pooled_db import PooledDB
from ai_analyzer import analyze_process_alert, test_ai_connection
from cache import init_redis, get_cache, set_cache, invalidate_cache
from cache_invalidation import on_client_update, on_alert_update, on_alert_settings_update
from collections import defaultdict
import threading

# 配置日志
logger = logging.getLogger('hwmon')
logger.setLevel(logging.INFO)
if not logger.handlers:
    handler = logging.StreamHandler()
    handler.setFormatter(logging.Formatter(
        '[%(asctime)s] %(levelname)s %(message)s',
        datefmt='%Y-%m-%d %H:%M:%S'
    ))
    logger.addHandler(handler)

app = Flask(__name__)

# CORS 配置 - 白名单模式
ALLOWED_ORIGINS = [
    "http://192.168.20.27:5000",
    "http://localhost:5000",
    "http://127.0.0.1:5000",
]

@app.before_request
def csrf_protect():
    """CSRF 防护 - 验证 Origin header"""
    # 只对写操作进行CSRF检查
    if request.method in ['POST', 'PUT', 'DELETE']:
        # 跳过不需要CSRF保护的端点（如客户端上报）
        exempt_paths = ['/api/report', '/api/process-alert', '/api/login']
        if request.path in exempt_paths:
            return

        origin = request.headers.get('Origin')
        # 如果有Origin头，验证是否在白名单中
        if origin and origin not in ALLOWED_ORIGINS:
            return jsonify({'error': 'CSRF validation failed'}), 403


@app.after_request
def add_security_headers(response):
    """添加安全响应头"""
    # CORS 白名单
    origin = request.headers.get('Origin')
    if origin in ALLOWED_ORIGINS:
        response.headers['Access-Control-Allow-Origin'] = origin
        response.headers['Access-Control-Allow-Credentials'] = 'true'
        response.headers['Access-Control-Allow-Headers'] = 'Content-Type, Authorization'
        response.headers['Access-Control-Allow-Methods'] = 'GET, POST, PUT, DELETE, OPTIONS'

    # 安全响应头
    response.headers['X-Frame-Options'] = 'DENY'
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['X-XSS-Protection'] = '1; mode=block'
    response.headers['Referrer-Policy'] = 'strict-origin-when-cross-origin'

    return response

# Session 安全配置
app.config['SESSION_COOKIE_SECURE'] = False  # 生产环境设为 True（需要 HTTPS）
app.config['SESSION_COOKIE_HTTPONLY'] = True  # 防止 XSS 访问 Cookie
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'  # 防止 CSRF
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(hours=24)  # Session 24小时过期

# 登录速率限制（基于IP）
login_attempts = defaultdict(list)  # {ip: [timestamp1, timestamp2, ...]}
LOGIN_RATE_LIMIT = 5  # 每分钟最多尝试次数
LOGIN_RATE_WINDOW = 60  # 时间窗口（秒）

def check_login_rate_limit(ip):
    """检查登录速率限制"""
    now = time.time()
    # 清理过期记录
    login_attempts[ip] = [t for t in login_attempts[ip] if now - t < LOGIN_RATE_WINDOW]

    if len(login_attempts[ip]) >= LOGIN_RATE_LIMIT:
        return False  # 超过限制

    login_attempts[ip].append(now)
    return True


# SSRF 防护
BLOCKED_HOSTS = ['169.254.169.254', '127.0.0.1', '0.0.0.0', 'localhost']

def validate_probe_host(host):
    """验证探针目标地址（防止SSRF）"""
    if host in BLOCKED_HOSTS:
        return False

    # 尝试解析为IP地址
    try:
        import ipaddress
        ip = ipaddress.ip_address(host)
        # 屏蔽私有网络和云元数据端点
        if ip.is_private or ip.is_loopback or ip.is_link_local:
            return False
    except ValueError:
        # 不是IP地址，可能是域名，允许通过
        pass

    return True

# 加载配置文件
import secrets

def load_config():
    """加载配置文件（从exe同级目录加载，不打包进exe）"""
    # 获取配置文件路径
    if getattr(sys, 'frozen', False):
        # PyInstaller打包后：从exe所在目录加载config.json
        base_path = os.path.dirname(sys.executable)
    else:
        # 开发环境的路径
        base_path = os.path.dirname(os.path.abspath(__file__))
    
    config_file = os.path.join(base_path, 'config.json')
    
    # 默认配置（仅作为fallback，不应包含真实密码）
    default_config = {
        'database': {
            'host': 'localhost',
            'port': 3306,
            'user': 'root',
            'password': '',  # 空密码，必须通过config.json或环境变量配置
            'database': 'hwmon',
            'charset': 'utf8mb4'
        },
        'login': {
            'username': 'admin',
            'password': ''  # 空密码，必须配置
        },
        'server': {
            'port': 5000,
            'host': '0.0.0.0'
        },
        'collect': {
            'max_workers': 50,
            'timeout': 15,
            'retry_times': 0
        }
    }
    
    try:
        with open(config_file, 'r', encoding='utf-8') as f:
            config = json.load(f)
            # 合并默认配置
            for key in default_config:
                if key not in config:
                    config[key] = default_config[key]
            return config
    except Exception as e:
        print(f"警告: 无法加载配置文件 {config_file}: {e}")
        print("使用默认配置（请注意：默认配置不包含有效密码，请创建config.json）")
        return default_config

CONFIG = load_config()

# Session密钥 - 优先从环境变量读取，否则生成随机密钥
app.secret_key = os.environ.get('SECRET_KEY') or CONFIG.get('secret_key') or secrets.token_hex(32)

# 登录配置 - 优先从环境变量读取
LOGIN_CONFIG = {
    'username': os.environ.get('LOGIN_USERNAME') or CONFIG.get('login', {}).get('username', 'admin'),
    'password': os.environ.get('LOGIN_PASSWORD') or CONFIG.get('login', {}).get('password', '')
}

# 检查是否配置了有效密码
if not LOGIN_CONFIG['password']:
    print("⚠️  警告: 未配置登录密码！请设置 LOGIN_PASSWORD 环境变量或在 config.json 中配置")

# MySQL数据库配置（优先从环境变量读取）
db_config = CONFIG.get('database', {})
MYSQL_CONFIG = {
    'host': os.environ.get('DB_HOST') or db_config.get('host', 'localhost'),
    'port': int(os.environ.get('DB_PORT') or db_config.get('port', 3306)),
    'user': os.environ.get('DB_USER') or db_config.get('user', 'root'),
    'password': os.environ.get('DB_PASSWORD') or db_config.get('password', ''),
    'database': os.environ.get('DB_NAME') or db_config.get('database', 'hwmon'),
    'charset': os.environ.get('DB_CHARSET') or db_config.get('charset', 'utf8mb4'),
    'cursorclass': pymysql.cursors.DictCursor,
    'autocommit': True,
    'init_command': "SET time_zone='+08:00', innodb_lock_wait_timeout=10",
    'connect_timeout': 10,
    'read_timeout': 30,
    'write_timeout': 30
}

# 检查数据库配置
if not MYSQL_CONFIG['password']:
    print("⚠️  警告: 未配置数据库密码！请设置 DB_PASSWORD 环境变量或在 config.json 中配置")

# 数据库连接池
db_pool = None

# 并发采集配置（从配置文件读取）
collect_config = CONFIG.get('collect', {})
COLLECT_CONFIG = {
    'max_workers': collect_config.get('max_workers', 50),
    'timeout': collect_config.get('timeout', 15),
    'retry_times': collect_config.get('retry_times', 0),
}

# 硬件变更检测线程池（异步执行，不阻塞主流程）
hw_detection_executor = ThreadPoolExecutor(
    max_workers=50,  # 50个并发检测线程（提升并发，避免上报集中时堆积）
    thread_name_prefix='hw_detect'
)

# 硬件检测任务有界信号量：防止极端情况下检测任务无界堆积导致内存上涨
_detection_semaphore = threading.BoundedSemaphore(200)

# AI 研判专用线程池（与硬件检测分离，避免 25s+ 的 AI 调用长时间占用检测线程）
ai_executor = ThreadPoolExecutor(
    max_workers=5,
    thread_name_prefix='ai_analyze'
)


def init_db_pool():
    """初始化数据库连接池（优化为支持1500+客户端）"""
    global db_pool
    db_pool = PooledDB(
        creator=pymysql,
        maxconnections=300,     # 支持300并发（1500客户端/5分钟均匀分布）
        mincached=30,           # 预创建30个空闲连接
        maxcached=80,           # 最多缓存80个空闲连接
        maxusage=300,           # 每个连接使用300次后回收（更频繁刷新）
        blocking=False,         # 连接池满时不阻塞，快速失败
        ping=0,                 # 关闭每次取连接前的 ping 探测，省一次往返（断连由 get_db 重试兜底）
        **MYSQL_CONFIG
    )
    print("MySQL连接池初始化成功（配置：max=300, min=30, cache=80）")
    print("数据库时区: 东八区 (北京时间)")


def get_db():
    """从连接池获取数据库连接"""
    max_retries = 3
    for attempt in range(max_retries):
        try:
            conn = db_pool.connection()
            # 设置会话时区为东八区（北京时间）
            cursor = conn.cursor()
            cursor.execute("SET time_zone='+08:00'")
            return conn
        except Exception as e:
            if attempt < max_retries - 1:
                print(f'[WARN] 获取数据库连接失败，重试 {attempt + 1}/{max_retries}: {e}')
                import time
                time.sleep(0.5)
            else:
                raise e


def get_db_readonly():
    """获取只读数据库连接（用于查询操作，避免锁冲突）

    说明：InnoDB 普通 SELECT 本身是快照读、不阻塞写，
    故不再额外执行 SET SESSION TRANSACTION READ ONLY（省一次往返）。
    """
    max_retries = 3
    for attempt in range(max_retries):
        try:
            conn = db_pool.connection()
            cursor = conn.cursor()
            cursor.execute("SET time_zone='+08:00'")
            return conn
        except Exception as e:
            if attempt < max_retries - 1:
                print(f'[WARN] 获取只读连接失败，重试 {attempt + 1}/{max_retries}: {e}')
                import time
                time.sleep(0.5)
            else:
                raise e


from contextlib import contextmanager

@contextmanager
def get_db_safe():
    """安全的数据库连接上下文管理器（自动关闭连接）"""
    conn = None
    try:
        conn = get_db()
        yield conn
    finally:
        if conn:
            try:
                conn.close()
            except:
                pass


@contextmanager
def get_db_readonly_safe():
    """安全的只读数据库连接上下文管理器（自动关闭连接）"""
    conn = None
    try:
        conn = get_db_readonly()
        yield conn
    finally:
        if conn:
            try:
                conn.close()
            except:
                pass


def init_tables():
    """初始化数据库表结构"""
    conn = get_db()
    try:
        cursor = conn.cursor()
        
        # 设置会话时区为东八区
        cursor.execute("SET time_zone='+08:00'")
        
        # 创建分组表
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS `groups` (
                id INT AUTO_INCREMENT PRIMARY KEY,
                name VARCHAR(255) NOT NULL UNIQUE,
                description TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        ''')

        # 创建客户端表
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS clients (
                id INT AUTO_INCREMENT PRIMARY KEY,
                client_id VARCHAR(255) NOT NULL UNIQUE,
                hostname VARCHAR(255),
                local_ip VARCHAR(45),
                group_id INT,
                last_report DATETIME,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (group_id) REFERENCES `groups`(id) ON DELETE SET NULL,
                INDEX idx_client_id (client_id),
                INDEX idx_group_id (group_id),
                INDEX idx_last_report (last_report),
                INDEX idx_last_report_desc (last_report DESC)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        ''')

        # 创建硬件信息历史表
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS hardware_reports (
                id INT AUTO_INCREMENT PRIMARY KEY,
                client_id VARCHAR(255) NOT NULL,
                report_data LONGTEXT,
                report_type VARCHAR(50) DEFAULT 'scheduled',
                timestamp DATETIME DEFAULT CURRENT_TIMESTAMP,
                INDEX idx_client_id (client_id),
                INDEX idx_timestamp (timestamp)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        ''')

        # 创建硬件采集历史表
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS hardware_history (
                id INT AUTO_INCREMENT PRIMARY KEY,
                client_id VARCHAR(255) NOT NULL,
                cpu_info TEXT,
                memory_info TEXT,
                disk_info TEXT,
                gpu_info TEXT,
                snapshot LONGTEXT,
                timestamp DATETIME DEFAULT CURRENT_TIMESTAMP,
                INDEX idx_client_id (client_id),
                INDEX idx_timestamp (timestamp),
                INDEX idx_client_timestamp (client_id, timestamp DESC)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        ''')

        # 创建默认分组
        cursor.execute('INSERT IGNORE INTO `groups` (name, description) VALUES (%s, %s)',
                       ('默认分组', '未分组的客户端'))

        # 创建客户端硬件基准表
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS client_baselines (
                client_id VARCHAR(255) PRIMARY KEY,
                cpu_snapshot TEXT,
                gpu_snapshot TEXT,
                memory_snapshot TEXT,
                disk_snapshot TEXT,
                baseline_timestamp TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (client_id) REFERENCES clients(client_id) ON DELETE CASCADE
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        ''')

        # 数据库迁移：client_baselines 增加「手动基准」字段
        #   selected_items   TEXT(JSON) 勾选纳入对比的条目 key 列表（按硬件类型分组）
        #   manual_overrides TEXT(JSON) 手动覆盖字段值（item_key -> {field: value}），优先级最高
        #   baseline_mode    VARCHAR     auto=自动全量 / manual=手动选择
        try:
            baseline_cols = [
                ('selected_items', "ALTER TABLE client_baselines ADD COLUMN selected_items TEXT"),
                ('manual_overrides', "ALTER TABLE client_baselines ADD COLUMN manual_overrides TEXT"),
                ('baseline_mode', "ALTER TABLE client_baselines ADD COLUMN baseline_mode VARCHAR(20) DEFAULT 'auto'"),
            ]
            for col_name, ddl in baseline_cols:
                cursor.execute(f"SHOW COLUMNS FROM client_baselines LIKE '{col_name}'")
                if not cursor.fetchone():
                    print(f'[INFO] 正在迁移数据库：添加 client_baselines.{col_name} 字段')
                    cursor.execute(ddl)
            conn.commit()
        except Exception as e:
            print(f'[WARN] client_baselines 手动基准字段迁移失败: {e}')

        # 创建 CPU 真实型号映射表（手动映射：占位 CPU 底层名 -> 真实型号）
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS cpu_name_map (
                client_id VARCHAR(255) PRIMARY KEY,
                cpu_real_name VARCHAR(255) NOT NULL,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                FOREIGN KEY (client_id) REFERENCES clients(client_id) ON DELETE CASCADE
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        ''')

        # 创建告警记录表
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS alert_records (
                id INT AUTO_INCREMENT PRIMARY KEY,
                client_id VARCHAR(255) NOT NULL,
                alert_type VARCHAR(100) NOT NULL,
                alert_detail LONGTEXT NOT NULL,
                resolved TINYINT DEFAULT 0,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                INDEX idx_client_id (client_id),
                INDEX idx_resolved (resolved),
                INDEX idx_created_at (created_at),
                INDEX idx_created_resolved (created_at DESC, resolved)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        ''')

        # 创建邮件配置表
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS email_config (
                id INT PRIMARY KEY,
                smtp_host VARCHAR(255) NOT NULL DEFAULT 'smtp.qq.com',
                smtp_port INT NOT NULL DEFAULT 465,
                smtp_user VARCHAR(255) NOT NULL DEFAULT '',
                smtp_password VARCHAR(255) NOT NULL DEFAULT '',
                sender_name VARCHAR(255) DEFAULT '硬件监控系统',
                recipients TEXT NOT NULL,
                enabled TINYINT DEFAULT 0
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        ''')
        cursor.execute('INSERT IGNORE INTO email_config (id, recipients) VALUES (1, "[]")')

        # 创建告警设置表
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS alert_settings (
                id INT PRIMARY KEY,
                alert_enabled TINYINT DEFAULT 1 COMMENT '告警开关：1=开启告警记录，0=关闭告警',
                email_enabled TINYINT DEFAULT 0 COMMENT '邮件开关：1=开启邮件通知，0=关闭邮件',
                monitor_cpu TINYINT DEFAULT 1,
                monitor_gpu TINYINT DEFAULT 1,
                monitor_memory TINYINT DEFAULT 1,
                monitor_disk TINYINT DEFAULT 1,
                monitor_network TINYINT DEFAULT 0,
                monitor_motherboard TINYINT DEFAULT 0,
                monitor_bios TINYINT DEFAULT 0,
                monitor_temperature TINYINT DEFAULT 0 COMMENT '监控温度传感器变化',
                monitor_fan TINYINT DEFAULT 0 COMMENT '监控风扇传感器变化',
                monitor_voltage TINYINT DEFAULT 0 COMMENT '监控电压传感器变化',
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        ''')
        cursor.execute('INSERT IGNORE INTO alert_settings (id) VALUES (1)')

        # 数据库迁移：为已存在的表添加新字段
        try:
            # 检查 alert_enabled 字段是否存在
            cursor.execute("SHOW COLUMNS FROM alert_settings LIKE 'alert_enabled'")
            if not cursor.fetchone():
                print('[INFO] 正在迁移数据库：添加 alert_enabled 字段')
                cursor.execute("ALTER TABLE alert_settings ADD COLUMN alert_enabled TINYINT DEFAULT 1 COMMENT '告警开关' AFTER id")
            
            # 检查 email_enabled 字段是否存在
            cursor.execute("SHOW COLUMNS FROM alert_settings LIKE 'email_enabled'")
            if not cursor.fetchone():
                print('[INFO] 正在迁移数据库：添加 email_enabled 字段')
                cursor.execute("ALTER TABLE alert_settings ADD COLUMN email_enabled TINYINT DEFAULT 0 COMMENT '邮件开关' AFTER alert_enabled")

            # 检查 monitor_temperature 字段是否存在
            cursor.execute("SHOW COLUMNS FROM alert_settings LIKE 'monitor_temperature'")
            if not cursor.fetchone():
                print('[INFO] 正在迁移数据库：添加 monitor_temperature 字段')
                cursor.execute("ALTER TABLE alert_settings ADD COLUMN monitor_temperature TINYINT DEFAULT 0 COMMENT '监控温度传感器变化'")

            # 检查 monitor_fan 字段是否存在
            cursor.execute("SHOW COLUMNS FROM alert_settings LIKE 'monitor_fan'")
            if not cursor.fetchone():
                print('[INFO] 正在迁移数据库：添加 monitor_fan 字段')
                cursor.execute("ALTER TABLE alert_settings ADD COLUMN monitor_fan TINYINT DEFAULT 0 COMMENT '监控风扇传感器变化'")

            # 检查 monitor_voltage 字段是否存在
            cursor.execute("SHOW COLUMNS FROM alert_settings LIKE 'monitor_voltage'")
            if not cursor.fetchone():
                print('[INFO] 正在迁移数据库：添加 monitor_voltage 字段')
                cursor.execute("ALTER TABLE alert_settings ADD COLUMN monitor_voltage TINYINT DEFAULT 0 COMMENT '监控电压传感器变化'")

            conn.commit()
            print('[INFO] 数据库字段迁移完成')
        except Exception as e:
            print(f'[WARN] 数据库字段迁移失败: {e}')
            # 不中断启动，继续运行

        # 数据库迁移：添加复合索引
        try:
            # 检查 hardware_history 表的复合索引
            cursor.execute("SHOW INDEX FROM hardware_history WHERE Key_name = 'idx_client_timestamp'")
            if not cursor.fetchone():
                print('[INFO] 正在添加硬件历史表复合索引')
                cursor.execute("ALTER TABLE hardware_history ADD INDEX idx_client_timestamp (client_id, timestamp DESC)")
            
            # 检查 clients 表的降序索引
            cursor.execute("SHOW INDEX FROM clients WHERE Key_name = 'idx_last_report_desc'")
            if not cursor.fetchone():
                print('[INFO] 正在添加客户端表降序索引')
                cursor.execute("ALTER TABLE clients ADD INDEX idx_last_report_desc (last_report DESC)")
            
            # 检查 alert_records 表的复合索引
            cursor.execute("SHOW INDEX FROM alert_records WHERE Key_name = 'idx_created_resolved'")
            if not cursor.fetchone():
                print('[INFO] 正在添加告警记录表复合索引')
                cursor.execute("ALTER TABLE alert_records ADD INDEX idx_created_resolved (created_at DESC, resolved)")
            
            # 检查 alert_records 表的去重查询复合索引（优化硬件变更去重）
            cursor.execute("SHOW INDEX FROM alert_records WHERE Key_name = 'idx_client_resolved_type'")
            if not cursor.fetchone():
                print('[INFO] 正在添加告警去重复合索引')
                cursor.execute("ALTER TABLE alert_records ADD INDEX idx_client_resolved_type (client_id, resolved, alert_type)")
            
            conn.commit()
            print('[INFO] 数据库索引迁移完成')
        except Exception as e:
            print(f'[WARN] 数据库索引迁移失败: {e}')

        # 创建进程告警记录表
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS process_alert_records (
                id INT AUTO_INCREMENT PRIMARY KEY,
                client_id VARCHAR(255) NOT NULL,
                hostname VARCHAR(255),
                local_ip VARCHAR(45),
                alert_data LONGTEXT NOT NULL COMMENT '完整告警 JSON',
                alert_count INT DEFAULT 1 COMMENT '本次告警包含的进程数量',
                resolved TINYINT DEFAULT 0,
                ai_analyzed TINYINT DEFAULT 0 COMMENT '是否已进行 AI 研判',
                ai_result LONGTEXT COMMENT 'AI 研判结果',
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                INDEX idx_pa_client_id (client_id),
                INDEX idx_pa_resolved (resolved),
                INDEX idx_pa_created_at (created_at DESC),
                INDEX idx_pa_ai_analyzed (ai_analyzed)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        ''')

        # 创建 AI 配置表
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS ai_config (
                id INT PRIMARY KEY,
                enabled TINYINT DEFAULT 0 COMMENT 'AI 研判总开关',
                api_base_url VARCHAR(500) DEFAULT 'https://api.openai.com/v1' COMMENT 'API 基础地址',
                api_key VARCHAR(500) DEFAULT '' COMMENT 'API Key',
                model VARCHAR(100) DEFAULT 'gpt-4o-mini' COMMENT '模型名称',
                max_tokens INT DEFAULT 2000 COMMENT '最大输出 token 数',
                temperature DECIMAL(3,2) DEFAULT 0.30 COMMENT '温度参数',
                system_prompt TEXT COMMENT '系统提示词',
                auto_analyze TINYINT DEFAULT 1 COMMENT '收到告警后自动触发 AI 分析',
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        ''')
        cursor.execute('INSERT IGNORE INTO ai_config (id) VALUES (1)')

        # 创建 AI 监控主机表
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS ai_monitored_hosts (
                id INT AUTO_INCREMENT PRIMARY KEY,
                client_id VARCHAR(255) NOT NULL UNIQUE,
                hostname VARCHAR(255),
                description TEXT,
                enabled TINYINT DEFAULT 1,
                added_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                INDEX idx_amh_client_id (client_id)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        ''')

        # 创建主机探测目标表
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS probe_targets (
                id INT AUTO_INCREMENT PRIMARY KEY,
                name VARCHAR(255) NOT NULL COMMENT '主机名称',
                host VARCHAR(255) NOT NULL COMMENT 'IP或域名',
                port INT DEFAULT 80 COMMENT '探测端口',
                protocol VARCHAR(10) DEFAULT 'http' COMMENT 'http/https/tcp',
                path VARCHAR(500) DEFAULT '/' COMMENT 'HTTP路径',
                enabled TINYINT DEFAULT 1,
                description TEXT,
                group_id INT COMMENT '分组ID',
                last_status VARCHAR(20) DEFAULT 'unknown' COMMENT 'unknown/online/offline',
                last_probe_time DATETIME,
                last_response_time INT COMMENT '响应时间ms',
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                INDEX idx_pt_enabled (enabled),
                INDEX idx_pt_group (group_id)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        ''')

        # 数据库迁移：为已存在的 probe_targets 表添加 group_id 字段
        try:
            cursor.execute("SHOW COLUMNS FROM probe_targets LIKE 'group_id'")
            if not cursor.fetchone():
                print('[INFO] 正在迁移数据库：添加 probe_targets.group_id 字段')
                cursor.execute("ALTER TABLE probe_targets ADD COLUMN group_id INT COMMENT '分组ID' AFTER description")
                cursor.execute("ALTER TABLE probe_targets ADD INDEX idx_pt_group (group_id)")
                conn.commit()
                print('[INFO] 数据库字段迁移完成')
        except Exception as e:
            print(f'[WARN] 数据库字段迁移失败: {e}')

        # 创建主机探测告警表
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS probe_alerts (
                id INT AUTO_INCREMENT PRIMARY KEY,
                target_id INT NOT NULL,
                target_name VARCHAR(255),
                target_host VARCHAR(255),
                alert_type VARCHAR(50) DEFAULT 'offline' COMMENT 'offline/timeout/slow',
                status_code INT COMMENT 'HTTP状态码',
                response_time INT COMMENT '响应时间ms',
                error_message TEXT,
                resolved TINYINT DEFAULT 0,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                INDEX idx_pa_target (target_id),
                INDEX idx_pa_resolved (resolved),
                INDEX idx_pa_created (created_at DESC),
                FOREIGN KEY (target_id) REFERENCES probe_targets(id) ON DELETE CASCADE
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        ''')

        conn.commit()
        print("数据库表初始化完成")
    except Exception as e:
        print(f"数据库初始化失败: {e}")
        raise
    finally:
        conn.close()



def login_required(f):
    """登录验证装饰器"""
    @functools.wraps(f)
    def decorated_function(*args, **kwargs):
        if not check_session_valid():
            # 如果是API请求，返回401
            if request.path.startswith('/api/'):
                return jsonify({'error': '未登录', 'need_login': True}), 401
            # 否则重定向到登录页面
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function


# ==================== 硬件变更检测与邮件告警 ====================

def _parse_json_field(value):
    """安全解析 JSON 字段（可能是 None / str / 已解析的 dict/list）。

    返回 dict/list，解析失败或空值返回 None。
    """
    if value is None or value == '':
        return None
    if isinstance(value, (dict, list)):
        return value
    if isinstance(value, str):
        try:
            parsed = json.loads(value)
            return parsed if isinstance(parsed, (dict, list)) else None
        except (json.JSONDecodeError, TypeError):
            return None
    return None


def compare_hardware(baseline_snapshots, new_hardware, alert_settings=None, manual_config=None):
    """
    智能硬件变更检测

    核心原则：可以占便宜但不能吃亏
    - 新增硬件 → 不告警（可能是驱动安装）
    - 丢失硬件 → 必须告警（可能是被盗）
    - 升级（6G→8G）→ 不告警
    - 降级（8G→6G）→ 告警

    baseline_snapshots: 基准快照 {'cpu': JSON, 'gpu': JSON, 'memory': JSON, 'disk': JSON}
    new_hardware: 新的硬件信息dict
    alert_settings: 告警设置dict
    manual_config: 手动基准配置 {'selected_items': {...}, 'manual_overrides': {...}}；
                   前者是勾选纳入的条目 key（按硬件类型分组），后者是手动字段覆盖。
                   None 表示自动全量（等价旧行为）。

    返回: {
        'changes': [...],           # 需要告警的变更
        'upgrades': [...],          # 升级（仅记录）
        'should_alert': bool        # 是否需要告警
    }
    """
    from compare_engine import compare_items, should_trigger_alert, apply_manual_baseline

    # 解析手动基准配置（缺省安全默认）
    selected_map = (manual_config or {}).get('selected_items') or {}
    override_map = (manual_config or {}).get('manual_overrides') or {}

    result = {
        'changes': [],        # 需要即时告警的变更（仅降级）
        'upgrades': [],       # 升级（仅记录）
        'empty_changes': [],  # 丢失待确认（空集合 + 数量减少，需连续计数才告警）
        'empty_types': [],    # 本次存在丢失待确认的硬件类型
        'should_alert': False
    }

    # 如果没有提供告警设置，使用默认设置
    if alert_settings is None:
        alert_settings = {
            'monitor_cpu': 1,
            'monitor_gpu': 1,
            'monitor_memory': 1,
            'monitor_disk': 1,
            'monitor_network': 0,
            'monitor_motherboard': 0,
            'monitor_bios': 0,
        }

    # 硬件类型映射
    hardware_types = [
        ('cpu', 'CPU', 'monitor_cpu'),
        ('gpu', 'GPU', 'monitor_gpu'),
        ('memory', '内存', 'monitor_memory'),
        ('disk', '硬盘', 'monitor_disk'),
    ]

    for hw_key, hw_label, setting_key in hardware_types:
        if not alert_settings.get(setting_key, 1):
            continue

        # 获取基准数据
        baseline_raw = baseline_snapshots.get(hw_key, '[]' if hw_key != 'memory' else '{}')
        try:
            baseline_items = json.loads(baseline_raw) if baseline_raw else []
        except (json.JSONDecodeError, TypeError):
            baseline_items = []

        # 内存特殊处理：转换为列表格式
        if hw_key == 'memory':
            if isinstance(baseline_items, dict):
                baseline_items = baseline_items.get('modules', [])
            new_mem = new_hardware.get('memory', {})
            new_items = new_mem.get('modules', []) if isinstance(new_mem, dict) else []
        else:
            new_items = new_hardware.get(hw_key, [])

        # 使用智能对比引擎
        if baseline_items or new_items:
            # 手动基准：先按勾选项过滤 + 手动值覆盖，再对比（优先级最高）
            # 注意：空列表是有效配置（该类型全部取消勾选=整体不监控），
            # 只有 None 才表示该类型未配置（=全量对比，向后兼容）
            hw_selected = selected_map.get(hw_key)
            hw_overrides = override_map.get(hw_key)
            if hw_selected is not None or hw_overrides:
                baseline_items = apply_manual_baseline(
                    baseline_items, hw_key,
                    selected_keys=hw_selected,
                    manual_overrides=hw_overrides if hw_overrides else None,
                )

            compare_result = compare_items(baseline_items, new_items, hw_key)

            # GPU 占位状态抑制丢失告警：本次上报存在占位显卡（=独显驱动
            # 尚未加载完成），此时 GPU 的「丢失」判定暂缓，避免把"驱动未就绪"
            # 误判成"独显被盗"。驱动加载、占位显卡消失后自动恢复对比。
            if hw_key == 'gpu' and compare_result.get('gpu_placeholder'):
                for item in compare_result.get('upgraded', []):
                    result['upgrades'].append({
                        'type': hw_key,
                        'label': hw_label,
                        'change_type': 'upgraded',
                        'item': item.get('item', ''),
                        'dimension': item.get('dimension', ''),
                        'old': item.get('old', ''),
                        'new': item.get('new', ''),
                        'message': item.get('message', f'{hw_label}升级')
                    })
                continue

            # 空采集（基准有但本次采集为空）：把丢失候选放入 empty_changes，
            # 交由上层按「连续 N 次空采集」策略决定是否告警（方案 A）
            if compare_result.get('empty'):
                result['empty_types'].append(hw_key)
                for item in compare_result.get('lost', []):
                    result['empty_changes'].append({
                        'type': hw_key,
                        'label': hw_label,
                        'change_type': 'lost',
                        'item': item.get('item', ''),
                        'count': item.get('count', 1),
                        'message': item.get('message', f'{hw_label}丢失')
                    })
                continue

            # 处理丢失（数量减少）—— 与空集合一样纳入「连续 N 次确认」，
            # 抑制采集抖动导致的 key 不匹配/偶发少采误报（根因 B）
            if compare_result.get('lost'):
                result['empty_types'].append(hw_key)
                for item in compare_result.get('lost', []):
                    result['empty_changes'].append({
                        'type': hw_key,
                        'label': hw_label,
                        'change_type': 'lost',
                        'item': item.get('item', ''),
                        'count': item.get('count', 1),
                        'message': item.get('message', f'{hw_label}丢失')
                    })

            # 处理降级（中重要性降值）
            for item in compare_result.get('downgraded', []):
                result['changes'].append({
                    'type': hw_key,
                    'label': hw_label,
                    'change_type': 'downgraded',
                    'item': item.get('item', ''),
                    'dimension': item.get('dimension', ''),
                    'old': item.get('old', ''),
                    'new': item.get('new', ''),
                    'message': item.get('message', f'{hw_label}降级')
                })

            # 处理升级（中重要性升值，仅记录）
            for item in compare_result.get('upgraded', []):
                result['upgrades'].append({
                    'type': hw_key,
                    'label': hw_label,
                    'change_type': 'upgraded',
                    'item': item.get('item', ''),
                    'dimension': item.get('dimension', ''),
                    'old': item.get('old', ''),
                    'new': item.get('new', ''),
                    'message': item.get('message', f'{hw_label}升级')
                })

    # 判断是否需要告警（丢失待确认也算，尽管最终由连续计数决定是否入库）
    result['should_alert'] = bool(result['changes']) or bool(result['empty_changes'])

    return result


def get_email_config(conn):
    """获取邮件配置"""
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM email_config WHERE id = 1')
    row = cursor.fetchone()
    if row:
        return dict(row)
    return None


def send_alert_email(client_id, hostname, local_ip, changes):
    """发送硬件变更告警邮件"""
    try:
        conn = get_db_readonly()  # 只读查询使用只读连接
        config = get_email_config(conn)
        conn.close()

        if not config or not config.get('enabled'):
            return False

        recipients = json.loads(config.get('recipients', '[]'))
        if not recipients:
            return False

        import smtplib
        from email.mime.text import MIMEText
        from email.mime.multipart import MIMEMultipart

        smtp_host = config['smtp_host']
        smtp_port = config['smtp_port']
        smtp_user = config['smtp_user']
        smtp_password = config['smtp_password']
        sender_name = config.get('sender_name', '硬件监控系统')

        # 构建邮件内容
        change_lines = []
        for c in changes:
            change_lines.append(f'<p><strong>{c["label"]}:</strong> {c["old"]} → {c["new"]}</p>')

        html_body = f'''
        <html><body style="font-family:Microsoft YaHei,Arial,sans-serif;">
        <h2 style="color:#e53e3e;">【硬件变更告警】</h2>
        <table style="border-collapse:collapse;">
            <tr><td style="padding:5px 10px;font-weight:bold;">客户端:</td><td style="padding:5px 10px;">{hostname or client_id}</td></tr>
            <tr><td style="padding:5px 10px;font-weight:bold;">IP地址:</td><td style="padding:5px 10px;">{local_ip or '-'}</td></tr>
            <tr><td style="padding:5px 10px;font-weight:bold;">变更时间:</td><td style="padding:5px 10px;">{datetime.now(TZ_CST).strftime("%Y-%m-%d %H:%M:%S")}</td></tr>
        </table>
        <h3 style="margin-top:20px;">变更详情:</h3>
        {''.join(change_lines)}
        <hr style="margin:20px 0;">
        <p style="color:#718096;">如需重置基准，请登录硬件监控系统进行操作。</p>
        </body></html>
        '''

        msg = MIMEMultipart('alternative')
        msg['Subject'] = f'【硬件变更告警】客户端 {hostname or client_id} 检测到硬件变更'
        msg['From'] = f'{sender_name} <{smtp_user}>'
        msg['To'] = ', '.join(recipients)
        msg.attach(MIMEText(html_body, 'html', 'utf-8'))

        # 发送邮件
        server = smtplib.SMTP_SSL(smtp_host, smtp_port)
        server.login(smtp_user, smtp_password)
        server.sendmail(smtp_user, recipients, msg.as_string())
        server.quit()
        return True

    except Exception as e:
        import traceback
        print(f'邮件发送失败: {str(e)}')
        print(traceback.format_exc())
        return False


# =================================================================
# 丢失连续计数（方案 A 扩展：抑制采集抖动导致的误报）
# 客户端每 120 秒上报一次。两类「丢失」不应立即告警，需连续确认：
#   1) 空集合：某类硬件采集偶发为空（如无独显机器的 Microsoft Basic
#      Display Adapter 时有时无）；
#   2) 数量减少：采集字段格式变化导致 baseline 与上报 key 不匹配
#      （如 CPU name 从 WMI 友好名变为底层名），被误判「丢了 N 个」。
# 仅当连续 EMPTY_THRESHOLD 次均丢失才判定并告警；任一次恢复即清零。
# =================================================================
EMPTY_THRESHOLD = 3
_empty_collect_state = {}      # {client_id: {hw_key: consecutive_loss_count}}
_empty_collect_lock = threading.Lock()


def _bump_empty_count(client_id, hw_key):
    """递增丢失计数，返回是否达到告警阈值。"""
    with _empty_collect_lock:
        client_state = _empty_collect_state.setdefault(client_id, {})
        client_state[hw_key] = client_state.get(hw_key, 0) + 1
        return client_state[hw_key] >= EMPTY_THRESHOLD


def _reset_empty_count(client_id, hw_key):
    """采集恢复正常，清零对应类型的丢失计数。"""
    with _empty_collect_lock:
        client_state = _empty_collect_state.get(client_id)
        if client_state:
            client_state[hw_key] = 0


def _check_hardware_changes(cursor, conn, client_id, hostname, local_ip, hardware_info,
                            cpu_info, gpu_info, mem_info, disk_info):
    """检查硬件变更（智能对比版本）"""
    try:
        # 检查是否有基准数据
        cursor.execute('SELECT * FROM client_baselines WHERE client_id = %s', (client_id,))
        baseline = cursor.fetchone()

        if not baseline:
            # 首次上报：自动创建基准
            cursor.execute('''
                INSERT INTO client_baselines (client_id, cpu_snapshot, gpu_snapshot, memory_snapshot, disk_snapshot)
                VALUES (%s, %s, %s, %s, %s)
            ''', (client_id, cpu_info, gpu_info, mem_info, disk_info))
            conn.commit()
            logger.info(f'客户端 {client_id} 首次上报，已自动创建基准')
        else:
            # 有基准：对比硬件变化
            # 获取告警设置
            cursor.execute('SELECT * FROM alert_settings WHERE id = 1')
            alert_settings_row = cursor.fetchone()
            alert_settings = dict(alert_settings_row) if alert_settings_row else None

            baseline_snapshots = {
                'cpu': baseline['cpu_snapshot'],
                'gpu': baseline['gpu_snapshot'],
                'memory': baseline['memory_snapshot'],
                'disk': baseline['disk_snapshot'],
            }

            # 读取手动基准配置（勾选项 + 手动覆盖），随 baseline 一起生效
            manual_config = None
            try:
                if baseline.get('baseline_mode') == 'manual':
                    manual_config = {
                        'selected_items': _parse_json_field(baseline.get('selected_items')),
                        'manual_overrides': _parse_json_field(baseline.get('manual_overrides')),
                    }
            except Exception as e:
                logger.warning(f'客户端 {client_id} 手动基准配置解析失败，回退全量对比: {e}')

            # 使用智能对比引擎
            compare_result = compare_hardware(baseline_snapshots, hardware_info, alert_settings, manual_config)

            # 记录升级信息（仅记录，不告警）
            if compare_result.get('upgrades'):
                for upgrade in compare_result['upgrades']:
                    logger.info(f'客户端 {client_id} 升级: {upgrade.get("message", "")}')

            # ---- 丢失连续计数（方案 A 扩展：覆盖空集合 + 数量减少）----
            # 降级（真值变化）直接进入告警候选
            final_changes = list(compare_result.get('changes', []))
            empty_types = compare_result.get('empty_types', [])
            empty_changes = compare_result.get('empty_changes', [])

            # 达阈值的丢失类型 → 加入告警候选；未达标仅累计不告警
            for hw_key in empty_types:
                if _bump_empty_count(client_id, hw_key):
                    for ec in empty_changes:
                        if ec.get('type') == hw_key:
                            final_changes.append(ec)
                    _reset_empty_count(client_id, hw_key)
                    logger.warning(f'客户端 {client_id} {hw_key} 连续丢失达阈值，判为丢失')

            # 本次无丢失的类型 → 恢复正常，清零计数
            for hw_key in ('cpu', 'gpu', 'memory', 'disk'):
                if hw_key not in empty_types:
                    _reset_empty_count(client_id, hw_key)

            # ---- 告警入库（含去重）----
            if final_changes:
                # 检查告警开关
                if alert_settings.get('alert_enabled', 1):
                    alert_detail = json.dumps(final_changes, ensure_ascii=False)

                    # 去重：同一客户端相同变更且未处理的告警不再重复入库
                    cursor.execute('''
                        SELECT id FROM alert_records
                        WHERE client_id = %s AND alert_type = 'hardware_change'
                          AND resolved = 0 AND alert_detail = %s
                        LIMIT 1
                    ''', (client_id, alert_detail))
                    dup = cursor.fetchone()

                    if dup:
                        logger.info(f'客户端 {client_id} 存在未处理的相同告警，跳过入库（去重）')
                    else:
                        cursor.execute('''
                            INSERT INTO alert_records (client_id, alert_type, alert_detail)
                            VALUES (%s, %s, %s)
                        ''', (client_id, 'hardware_change', alert_detail))
                        conn.commit()
                        logger.warning(f'客户端 {client_id} 检测到硬件变更: {len(final_changes)} 项')
                else:
                    logger.info(f'客户端 {client_id} 检测到硬件变更，但告警已关闭，未记录')

                # 检查邮件开关（独立于告警开关）
                if alert_settings.get('email_enabled', 0):
                    try:
                        email_sent = send_alert_email(client_id, hostname, local_ip, final_changes)
                        if email_sent:
                            logger.info(f'已向管理员发送告警邮件')
                        else:
                            logger.warning(f'告警邮件发送失败（可能未配置或配置错误）')
                    except Exception as e:
                        logger.warning(f'发送邮件异常: {e}')
                else:
                    logger.info(f'邮件通知已关闭，未发送邮件')
    except Exception as e:
        # 回滚事务
        try:
            conn.rollback()
        except:
            pass
        raise e


def _check_hardware_changes_async(client_id, hostname, local_ip, hardware_info,
                                   cpu_info, gpu_info, mem_info, disk_info):
    """异步硬件变更检测（独立线程，自己管理数据库连接）"""
    conn = None
    try:
        conn = get_db()
        cursor = conn.cursor()
        
        # 复用原有逻辑
        _check_hardware_changes(cursor, conn, client_id, hostname, local_ip, hardware_info,
                                cpu_info, gpu_info, mem_info, disk_info)
    except Exception as e:
        print(f'[WARN] 异步硬件检测失败 {client_id}: {e}')
    finally:
        if conn:
            try:
                conn.close()
            except:
                pass
        _detection_semaphore.release()


# =================================================================


def collect_single_client(client_id, local_ip):
    """采集单个客户端(用于并发执行) - 使用全局配置"""
    return collect_single_client_with_config(client_id, local_ip, COLLECT_CONFIG)


def collect_single_client_with_config(client_id, local_ip, config):
    """采集单个客户端(使用指定配置，避免全局状态污染)"""
    if not local_ip:
        return {
            'client_id': client_id,
            'status': 'unknown_ip',
            'message': 'IP地址未知'
        }

    try:
        response = requests.post(
            f'http://{local_ip}:13301/api/collect',
            json={'trigger': 'server'},
            timeout=config['timeout']
        )

        if response.status_code == 200:
            return {
                'client_id': client_id,
                'status': 'success',
                'message': '采集成功'
            }
        else:
            return {
                'client_id': client_id,
                'status': 'failed',
                'message': f'HTTP {response.status_code}'
            }

    except requests.exceptions.Timeout:
        return {
            'client_id': client_id,
            'status': 'timeout',
            'message': '连接超时'
        }
    except requests.exceptions.ConnectionError:
        return {
            'client_id': client_id,
            'status': 'offline',
            'message': '无法连接,客户端可能离线或防火墙阻止'
        }
    except Exception as e:
        return {
            'client_id': client_id,
            'status': 'error',
            'message': str(e)
        }


@app.before_request
def _ensure_tables():
    """确保数据库表存在（首次请求时自动创建）"""
    if not hasattr(app, '_tables_initialized'):
        try:
            init_tables()
        except Exception:
            pass
        app._tables_initialized = True


@app.route('/')
def index():
    """主页"""
    if 'logged_in' not in session:
        return redirect(url_for('login'))
    return render_template('index.html')


@app.route('/login')
def login():
    """登录页面"""
    return render_template('login.html')


@app.route('/api/login', methods=['POST'])
def api_login():
    """登录API"""
    try:
        # 检查速率限制
        client_ip = request.remote_addr
        if not check_login_rate_limit(client_ip):
            return jsonify({'status': 'error', 'message': '登录尝试过于频繁，请稍后再试'}), 429

        data = request.json
        username = data.get('username', '')
        password = data.get('password', '')

        if not username or not password:
            return jsonify({'status': 'error', 'message': '用户名和密码不能为空'}), 400

        # 验证用户名
        if username != LOGIN_CONFIG['username']:
            return jsonify({'status': 'error', 'message': '用户名或密码错误'}), 401

        # 验证密码（支持哈希和明文两种方式）
        from auth import verify_password, migrate_password_to_hash
        stored_hash = LOGIN_CONFIG.get('password_hash')
        stored_salt = LOGIN_CONFIG.get('password_salt')

        if stored_hash and stored_salt:
            # 使用哈希验证
            if not verify_password(password, stored_hash, stored_salt):
                return jsonify({'status': 'error', 'message': '用户名或密码错误'}), 401
        else:
            # 旧版明文密码验证
            if password != LOGIN_CONFIG['password']:
                return jsonify({'status': 'error', 'message': '用户名或密码错误'}), 401
            # 自动迁移到哈希
            try:
                new_hash, new_salt = migrate_password_to_hash(password)
                _migrate_password_to_hash(new_hash, new_salt)
            except Exception as e:
                logger.warning(f'密码迁移失败: {e}')

        # 生成唯一的 session ID
        import uuid
        session['session_id'] = str(uuid.uuid4())
        session['logged_in'] = True
        session['username'] = username
        session.permanent = True  # 设置 session 永久性
        logger.info(f'用户 {username} 登录成功，session_id: {session["session_id"][:8]}...')
        return jsonify({'status': 'success', 'message': '登录成功'})

    except Exception as e:
        return jsonify({'error': str(e)}), 500


def _migrate_password_to_hash(password_hash, password_salt):
    """将密码哈希保存到配置文件"""
    try:
        config_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'config.json')
        if os.path.exists(config_file):
            with open(config_file, 'r', encoding='utf-8') as f:
                config = json.load(f)
            config.setdefault('login', {})['password_hash'] = password_hash
            config.setdefault('login', {})['password_salt'] = password_salt
            with open(config_file, 'w', encoding='utf-8') as f:
                json.dump(config, f, indent=4, ensure_ascii=False)
            logger.info('密码哈希已保存到配置文件')
    except Exception as e:
        logger.error(f'保存密码哈希失败: {e}')


# 已登出的session黑名单（使用Redis或内存）
logged_out_sessions = set()

@app.route('/api/logout', methods=['POST'])
def api_logout():
    """登出API"""
    # 将当前session加入黑名单
    if 'logged_in' in session:
        session_id = session.get('session_id', '')
        if session_id:
            logged_out_sessions.add(session_id)
            logger.info(f'Session {session_id} 已加入黑名单')
    session.clear()
    return jsonify({'status': 'success', 'message': '已登出'})


def check_session_valid():
    """检查session是否有效（未被登出）"""
    if 'logged_in' not in session:
        return False
    session_id = session.get('session_id', '')
    if session_id in logged_out_sessions:
        return False
    return True


@app.route('/api/check-login', methods=['GET'])
def check_login():
    """检查登录状态"""
    if check_session_valid():
        return jsonify({'status': 'success', 'logged_in': True, 'username': session.get('username')})
    else:
        return jsonify({'status': 'success', 'logged_in': False})


@app.route('/api/report', methods=['POST'])
def receive_report():
    """接收客户端上报的硬件信息（同步处理）"""
    try:
        data = request.json
        client_id = data.get('client_id')
        hostname = data.get('hostname')
        hardware_info = data.get('hardware_info')
        local_ip = data.get('local_ip', '')

        if not client_id:
            return jsonify({'error': '缺少client_id'}), 400

        # 直接同步处理，确保数据被正确保存
        return _process_report(1, 3)

    except Exception as e:
        logger.error(f'接收上报失败: {e}')
        return jsonify({'error': str(e)}), 500


def _process_report(attempt, max_retries):
    """处理客户端上报数据（内部函数）"""
    conn = None
    try:
        data = request.json
        client_id = data.get('client_id')
        hostname = data.get('hostname')
        hardware_info = data.get('hardware_info')
        local_ip = data.get('local_ip', '')
        # APIPA/无效地址兜底：169.254.x.x（DHCP 失败自动地址）、回环等改用真实来源 IP
        if not local_ip or local_ip.startswith('169.254.') or local_ip in ('127.0.0.1', '0.0.0.0', '::1'):
            local_ip = request.remote_addr or ''
        report_type = data.get('report_type', 'scheduled')

        if not client_id:
            return jsonify({'error': '缺少client_id'}), 400

        conn = get_db()
        cursor = conn.cursor()

        # 步骤1: 快速更新客户端信息（单独事务，减少锁持有时间）
        current_time = datetime.now(TZ_CST).strftime('%Y-%m-%d %H:%M:%S')
        cursor.execute('''
            INSERT INTO clients (client_id, hostname, local_ip, last_report)
            VALUES (%s, %s, %s, %s)
            ON DUPLICATE KEY UPDATE
                hostname = VALUES(hostname),
                local_ip = VALUES(local_ip),
                last_report = VALUES(last_report)
        ''', (client_id, hostname, local_ip, current_time))
        conn.commit()  # 立即提交，释放锁

        # 步骤2: 保存硬件报告（单独事务）
        cursor.execute('''
            INSERT INTO hardware_reports (client_id, report_data, report_type)
            VALUES (%s, %s, %s)
        ''', (client_id, json.dumps(hardware_info, ensure_ascii=False), report_type))

        # 提取关键硬件指标
        cpu_info = json.dumps(hardware_info.get('cpu', []), ensure_ascii=False) if hardware_info.get('cpu') else ''
        mem_info = json.dumps(hardware_info.get('memory', {}), ensure_ascii=False) if hardware_info.get('memory') else ''
        disk_info = json.dumps(hardware_info.get('disk', []), ensure_ascii=False) if hardware_info.get('disk') else ''
        gpu_info = json.dumps(hardware_info.get('gpu', []), ensure_ascii=False) if hardware_info.get('gpu') else ''

        cursor.execute('''
            INSERT INTO hardware_history (client_id, cpu_info, memory_info, disk_info, gpu_info, snapshot)
            VALUES (%s, %s, %s, %s, %s, %s)
        ''', (client_id, cpu_info, mem_info, disk_info, gpu_info, json.dumps(hardware_info, ensure_ascii=False)))

        # 【优化】移除实时DELETE，改为后台定时任务清理（见下方 cleanup_old_records 函数）
        # 这样每次上报不会执行耗时的DELETE操作，大幅提升性能
        conn.commit()  # 提交硬件历史相关操作

        conn.close()
        conn = None  # 标记已关闭

        # 【优化】步骤3: 硬件变更检测（异步执行，有界队列，不阻塞主流程）
        # 提交到线程池异步执行，立即返回，不等待结果
        if _detection_semaphore.acquire(blocking=False):
            try:
                hw_detection_executor.submit(
                    _check_hardware_changes_async,
                    client_id, hostname, local_ip, hardware_info,
                    cpu_info, gpu_info, mem_info, disk_info
                )
            except Exception:
                _detection_semaphore.release()
        else:
            logger.warning(f'硬件检测队列已满，跳过本次检测 {client_id}')

        return jsonify({'status': 'success', 'message': '接收成功'})

    except Exception as e:
        raise e  # 抛出异常，由外层重试机制处理


@app.route('/api/clients', methods=['GET'])
@login_required
def get_clients():
    """获取所有客户端列表（支持排序、分组过滤和关键词搜索）"""
    try:
        group_id = request.args.get('group_id')
        sort_by = request.args.get('sort_by', 'last_report')  # 默认按最后上报时间排序
        order = request.args.get('order', 'desc')  # 默认降序
        search = (request.args.get('search') or '').strip()  # 关键词搜索（主机名/ID/IP）

        # 验证排序字段（提前到缓存键生成前，保证缓存键稳定）
        valid_sort_fields = ['hostname', 'local_ip', 'group_name', 'last_report', 'created_at']
        if sort_by not in valid_sort_fields:
            sort_by = 'last_report'
        # 验证排序方向
        order = 'DESC' if order.lower() == 'desc' else 'ASC'

        # 短缓存（5s）：客户端列表是高频读取热路径，命中直接返回，绕过 DB 查询。
        # 注意：带搜索关键词时不走缓存（关键词组合千变万化，缓存命中率极低且会污染缓存）。
        if not search:
            cache_key = f'hwmon:clients:{group_id or "all"}:{sort_by}:{order}'
            cached = get_cache(cache_key)
            if cached:
                return jsonify(cached)
        else:
            cache_key = None

        with get_db_readonly_safe() as conn:
            cursor = conn.cursor()

            # 构建查询
            search_cond = ''
            search_params = []
            if search:
                like = f'%{search}%'
                search_cond = ' AND (c.hostname LIKE %s OR c.client_id LIKE %s OR c.local_ip LIKE %s)'
                search_params = [like, like, like]

            if group_id == 'ungrouped':
                cursor.execute(f'''
                    SELECT c.*, g.name as group_name
                    FROM clients c
                    LEFT JOIN `groups` g ON c.group_id = g.id
                    WHERE c.group_id IS NULL{search_cond}
                    ORDER BY c.{sort_by} {order}
                ''', search_params)
            elif group_id:
                cursor.execute(f'''
                    SELECT c.*, g.name as group_name
                    FROM clients c
                    LEFT JOIN `groups` g ON c.group_id = g.id
                    WHERE c.group_id = %s{search_cond}
                    ORDER BY c.{sort_by} {order}
                ''', [group_id] + search_params)
            else:
                cursor.execute(f'''
                    SELECT c.*, g.name as group_name
                    FROM clients c
                    LEFT JOIN `groups` g ON c.group_id = g.id
                    WHERE 1=1{search_cond}
                    ORDER BY c.{sort_by} {order}
                ''', search_params)

            clients = [dict(row) for row in cursor.fetchall()]

        response_data = {'status': 'success', 'data': clients}
        if cache_key:
            set_cache(cache_key, response_data, ttl=5)
        return jsonify(response_data)

    except Exception as e:
        logger.error(f'获取客户端列表失败: {e}')
        return jsonify({'error': str(e)}), 500


@app.route('/api/client/<client_id>', methods=['GET'])
@login_required
def get_client_detail(client_id):
    """获取客户端详细信息"""
    try:
        with get_db_safe() as conn:
            cursor = conn.cursor()

            # 获取客户端基本信息
            cursor.execute('''
                SELECT c.*, g.name as group_name
                FROM clients c
                LEFT JOIN `groups` g ON c.group_id = g.id
                WHERE c.client_id = %s
            ''', (client_id,))

            client = cursor.fetchone()
            if not client:
                return jsonify({'error': '客户端不存在'}), 404

            client_info = dict(client)

            # 获取 CPU 真实型号映射（手动映射，占位 CPU 名 -> 真实型号）
            cursor.execute('SELECT cpu_real_name FROM cpu_name_map WHERE client_id = %s', (client_id,))
            cpu_map = cursor.fetchone()
            if cpu_map:
                client_info['cpu_real_name'] = cpu_map['cpu_real_name']

            # 获取最新的硬件报告
            cursor.execute('''
                SELECT report_data, report_type, timestamp
                FROM hardware_reports
                WHERE client_id = %s
                ORDER BY timestamp DESC
                LIMIT 1
            ''', (client_id,))

            report = cursor.fetchone()
            if report:
                client_info['latest_hardware'] = json.loads(report['report_data'])
                client_info['last_hardware_update'] = report['timestamp']
                client_info['last_report_type'] = report['report_type']

        return jsonify({'status': 'success', 'data': client_info})

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/cpu-names', methods=['GET'])
@login_required
def get_cpu_name_maps():
    """获取所有 CPU 真实型号映射 {client_id: cpu_real_name}（手动映射）"""
    try:
        with get_db_readonly_safe() as conn:
            cursor = conn.cursor()
            cursor.execute('SELECT client_id, cpu_real_name FROM cpu_name_map')
            maps = {row['client_id']: row['cpu_real_name'] for row in cursor.fetchall()}
        return jsonify({'status': 'success', 'data': maps})
    except Exception as e:
        logger.error(f'获取 CPU 型号映射失败: {e}')
        return jsonify({'error': str(e)}), 500


@app.route('/api/client/<client_id>/cpu-name', methods=['POST'])
@login_required
def set_cpu_real_name(client_id):
    """录入/更新某台机器的 CPU 真实型号（手动映射）"""
    try:
        data = request.json or {}
        cpu_real_name = (data.get('cpu_real_name') or '').strip()
        if not cpu_real_name:
            return jsonify({'error': 'CPU 型号不能为空'}), 400

        with get_db_safe() as conn:
            cursor = conn.cursor()
            cursor.execute('SELECT client_id FROM clients WHERE client_id = %s', (client_id,))
            if not cursor.fetchone():
                return jsonify({'error': '客户端不存在'}), 404

            cursor.execute('''
                INSERT INTO cpu_name_map (client_id, cpu_real_name)
                VALUES (%s, %s)
                ON DUPLICATE KEY UPDATE cpu_real_name = VALUES(cpu_real_name)
            ''', (client_id, cpu_real_name))
            conn.commit()

        return jsonify({'status': 'success', 'message': 'CPU 真实型号已保存'})

    except Exception as e:
        logger.error(f'保存 CPU 型号映射失败: {e}')
        return jsonify({'error': str(e)}), 500


@app.route('/api/client/<client_id>/cpu-name', methods=['DELETE'])
@login_required
def delete_cpu_real_name(client_id):
    """删除某台机器的 CPU 真实型号映射"""
    try:
        with get_db_safe() as conn:
            cursor = conn.cursor()
            cursor.execute('DELETE FROM cpu_name_map WHERE client_id = %s', (client_id,))
            conn.commit()
        return jsonify({'status': 'success'})
    except Exception as e:
        logger.error(f'删除 CPU 型号映射失败: {e}')
        return jsonify({'error': str(e)}), 500


@app.route('/api/collect/<client_id>', methods=['POST'])
@login_required
def collect_from_client(client_id):
    """主动采集单个客户端"""
    try:
        conn = get_db_readonly()  # 只读查询使用只读连接
        cursor = conn.cursor()

        # 获取客户端信息
        cursor.execute('SELECT local_ip FROM clients WHERE client_id = %s', (client_id,))
        client = cursor.fetchone()

        if not client:
            conn.close()
            return jsonify({'error': '客户端不存在'}), 404

        local_ip = client['local_ip']
        conn.close()

        if not local_ip:
            return jsonify({'error': '客户端IP地址未知,无法采集'}), 400

        # 使用并发函数采集单个客户端
        result = collect_single_client(client_id, local_ip)

        if result['status'] == 'success':
            return jsonify({
                'status': 'success',
                'message': f'已向客户端 {client_id} 发送采集请求',
                'client_response': result
            })
        else:
            status_code = 500
            if result['status'] == 'timeout':
                status_code = 408
            elif result['status'] in ('offline', 'unknown_ip'):
                status_code = 404

            return jsonify({
                'status': result['status'],
                'message': result['message']
            }), status_code

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/collect/all', methods=['POST'])
@login_required
def collect_all_clients():
    """一键采集: 并发向所有客户端发送采集请求"""
    try:
        # 获取采集参数(支持自定义并发数和超时)
        params = request.json or {}
        max_workers = params.get('max_workers', COLLECT_CONFIG['max_workers'])
        timeout = params.get('timeout', COLLECT_CONFIG['timeout'])

        # 验证参数范围
        max_workers = max(1, min(max_workers, 200))  # 限制在1-200之间
        timeout = max(5, min(timeout, 60))  # 限制在5-60秒之间

        with get_db_readonly_safe() as conn:
            cursor = conn.cursor()

            cursor.execute('SELECT client_id, local_ip FROM clients')
            clients = [dict(row) for row in cursor.fetchall()]

        if not clients:
            return jsonify({
                'status': 'completed',
                'total': 0,
                'success': 0,
                'failed': 0,
                'results': [],
                'elapsed_seconds': 0
            })

        start_time = time.time()
        results = []

        # 创建局部配置副本（避免修改全局配置）
        local_collect_config = {
            'max_workers': max_workers,
            'timeout': timeout,
            'retry_times': COLLECT_CONFIG['retry_times']
        }

        # 使用线程池并发采集
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            # 提交所有采集任务
            futures = {
                executor.submit(collect_single_client_with_config, c['client_id'], c['local_ip'], local_collect_config): c['client_id']
                for c in clients
            }

            # 收集结果
            for future in as_completed(futures):
                try:
                    result = future.result()
                    results.append(result)
                except Exception as e:
                    results.append({
                        'client_id': futures[future],
                        'status': 'error',
                        'message': str(e)
                    })

        elapsed = time.time() - start_time

        # 统计结果
        success_count = sum(1 for r in results if r['status'] == 'success')
        fail_count = len(results) - success_count

        return jsonify({
            'status': 'completed',
            'total': len(clients),
            'success': success_count,
            'failed': fail_count,
            'elapsed_seconds': round(elapsed, 2),
            'concurrency': max_workers,
            'results': results
        })

    except Exception as e:
        logger.error(f'一键采集失败: {e}')
        return jsonify({'error': str(e)}), 500


@app.route('/api/groups', methods=['GET'])
@login_required
def get_groups():
    """获取所有分组（含全部主机总数与未分组数量，供侧栏「全部/未分组」计数使用）"""
    try:
        with get_db_readonly_safe() as conn:
            cursor = conn.cursor()

            cursor.execute('''
                SELECT g.*, COUNT(c.id) as client_count
                FROM `groups` g
                LEFT JOIN clients c ON g.id = c.group_id
                GROUP BY g.id
                ORDER BY g.name
            ''')

            groups = [dict(row) for row in cursor.fetchall()]

            # 全部主机总数（含未分组）；未分组的机器不属于任何分组，
            # 不能用各分组 client_count 求和代替（会漏掉未分组机器）
            cursor.execute('SELECT COUNT(*) as total FROM clients')
            total_clients = cursor.fetchone()['total']
            cursor.execute('SELECT COUNT(*) as cnt FROM clients WHERE group_id IS NULL')
            ungrouped_count = cursor.fetchone()['cnt']

        return jsonify({'status': 'success', 'data': groups,
                        'total_clients': total_clients,
                        'ungrouped_count': ungrouped_count})

    except Exception as e:
        logger.error(f'获取分组列表失败: {e}')
        return jsonify({'error': str(e)}), 500


@app.route('/api/groups', methods=['POST'])
@login_required
def create_group():
    """创建新分组"""
    try:
        data = request.json
        name = data.get('name')
        description = data.get('description', '')

        if not name:
            return jsonify({'error': '分组名称不能为空'}), 400

        with get_db_safe() as conn:
            cursor = conn.cursor()

            cursor.execute('INSERT INTO `groups` (name, description) VALUES (%s, %s)',
                           (name, description))

            conn.commit()
            group_id = cursor.lastrowid

        return jsonify({'status': 'success', 'group_id': group_id})

    except pymysql.err.IntegrityError:
        return jsonify({'error': '分组名称已存在'}), 400
    except Exception as e:
        logger.error(f'创建分组失败: {e}')
        return jsonify({'error': str(e)}), 500


@app.route('/api/groups/<int:group_id>', methods=['PUT'])
@login_required
def update_group(group_id):
    """更新分组"""
    try:
        data = request.json
        name = data.get('name')
        description = data.get('description')

        with get_db_safe() as conn:
            cursor = conn.cursor()

            cursor.execute('UPDATE `groups` SET name = %s, description = %s WHERE id = %s',
                           (name, description, group_id))

            conn.commit()

        return jsonify({'status': 'success'})

    except Exception as e:
        logger.error(f'更新分组失败: {e}')
        return jsonify({'error': str(e)}), 500


@app.route('/api/groups/<int:group_id>', methods=['DELETE'])
@login_required
def delete_group(group_id):
    """删除分组"""
    try:
        with get_db_safe() as conn:
            cursor = conn.cursor()

            # 检查是否是默认分组
            cursor.execute('SELECT name FROM `groups` WHERE id = %s', (group_id,))
            group = cursor.fetchone()
            if group and group['name'] == '默认分组':
                return jsonify({'error': '不能删除默认分组'}), 400

            cursor.execute('DELETE FROM `groups` WHERE id = %s', (group_id,))
            conn.commit()

        return jsonify({'status': 'success'})

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/clients/<client_id>/group', methods=['PUT'])
@login_required
def assign_client_to_group(client_id):
    """将客户端分配到分组"""
    try:
        data = request.json
        group_id = data.get('group_id')

        with get_db_safe() as conn:
            cursor = conn.cursor()

            cursor.execute('UPDATE clients SET group_id = %s WHERE client_id = %s',
                           (group_id, client_id))

            conn.commit()

        return jsonify({'status': 'success'})

    except Exception as e:
        logger.error(f'分配客户端分组失败: {e}')
        return jsonify({'error': str(e)}), 500


# 注意：批量分组接口 PUT /api/clients/batch-group 已在文件后部（export_excel 之后）定义，
# 此处不再重复定义，避免 Flask 路由 endpoint 冲突导致启动失败。


@app.route('/api/clients/<client_id>', methods=['DELETE'])
@login_required
def delete_client(client_id):
    """删除客户端"""
    try:
        with get_db_safe() as conn:
            cursor = conn.cursor()

            cursor.execute('DELETE FROM clients WHERE client_id = %s', (client_id,))
            conn.commit()

        return jsonify({'status': 'success'})

    except Exception as e:
        logger.error(f'删除客户端失败: {e}')
        return jsonify({'error': str(e)}), 500


@app.route('/api/export/csv', methods=['GET'])
@login_required
def export_csv():
    """导出所有客户端信息为CSV"""
    try:
        group_id = request.args.get('group_id')
        with get_db_readonly_safe() as conn:
            cursor = conn.cursor()

            if group_id == 'ungrouped':
                cursor.execute('''
                    SELECT c.client_id, c.hostname, c.local_ip, g.name as group_name,
                           c.last_report, c.created_at
                    FROM clients c
                    LEFT JOIN `groups` g ON c.group_id = g.id
                    WHERE c.group_id IS NULL
                    ORDER BY c.last_report DESC
                ''')
            elif group_id:
                cursor.execute('''
                    SELECT c.client_id, c.hostname, c.local_ip, g.name as group_name,
                           c.last_report, c.created_at
                    FROM clients c
                    LEFT JOIN `groups` g ON c.group_id = g.id
                    WHERE c.group_id = %s
                    ORDER BY c.last_report DESC
                ''', (group_id,))
            else:
                cursor.execute('''
                    SELECT c.client_id, c.hostname, c.local_ip, g.name as group_name,
                           c.last_report, c.created_at
                    FROM clients c
                    LEFT JOIN `groups` g ON c.group_id = g.id
                    ORDER BY c.last_report DESC
                ''')

            clients = cursor.fetchall()

        # 创建CSV
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(['客户端ID(主机名)', 'IP地址', '分组', '最后上报时间', '创建时间'])

        for client in clients:
            writer.writerow([
                client['client_id'],
                client['local_ip'] or '-',
                client['group_name'] or '未分组',
                client['last_report'],
                client['created_at']
            ])

        output.seek(0)
        return send_file(
            io.BytesIO(output.getvalue().encode('utf-8-sig')),
            mimetype='text/csv',
            as_attachment=True,
            download_name=f'hardware_report_{datetime.now(TZ_CST).strftime("%Y%m%d_%H%M%S")}.csv'
        )

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/export/json', methods=['GET'])
@login_required
def export_json():
    """导出所有客户端信息为JSON"""
    try:
        def json_serializer(obj):
            """自定义JSON序列化器，处理datetime类型"""
            if hasattr(obj, 'isoformat'):
                return obj.isoformat()
            if hasattr(obj, 'strftime'):
                return obj.strftime('%Y-%m-%d %H:%M:%S')
            return str(obj)

        group_id = request.args.get('group_id')
        with get_db_readonly_safe() as conn:
            cursor = conn.cursor()

            if group_id == 'ungrouped':
                cursor.execute('''
                    SELECT c.*, g.name as group_name
                    FROM clients c
                    LEFT JOIN `groups` g ON c.group_id = g.id
                    WHERE c.group_id IS NULL
                    ORDER BY c.last_report DESC
                ''')
            elif group_id:
                cursor.execute('''
                    SELECT c.*, g.name as group_name
                    FROM clients c
                    LEFT JOIN `groups` g ON c.group_id = g.id
                    WHERE c.group_id = %s
                    ORDER BY c.last_report DESC
                ''', (group_id,))
            else:
                cursor.execute('''
                    SELECT c.*, g.name as group_name
                    FROM clients c
                    LEFT JOIN `groups` g ON c.group_id = g.id
                    ORDER BY c.last_report DESC
                ''')

            clients = [dict(row) for row in cursor.fetchall()]

            # 获取每个客户端的最新硬件信息（N+1 问题，后续优化）
            for client in clients:
                cursor.execute('''
                    SELECT report_data, report_type, timestamp
                    FROM hardware_reports
                    WHERE client_id = %s
                    ORDER BY timestamp DESC
                    LIMIT 1
                ''', (client['client_id'],))

                report = cursor.fetchone()
                if report:
                    client['latest_hardware'] = json.loads(report['report_data'])
                    client['last_hardware_update'] = report['timestamp']
                    client['last_report_type'] = report['report_type']

        return send_file(
            io.BytesIO(json.dumps(clients, ensure_ascii=False, indent=2, default=json_serializer).encode('utf-8')),
            mimetype='application/json',
            as_attachment=True,
            download_name=f'hardware_report_{datetime.now(TZ_CST).strftime("%Y%m%d_%H%M%S")}.json'
        )

    except Exception as e:
        logger.error(f'导出JSON失败: {e}')
        return jsonify({'error': str(e)}), 500


@app.route('/api/export/excel', methods=['GET'])
@login_required
def export_excel():
    """导出客户端硬件信息为Excel文件"""
    try:
        group_id = request.args.get('group_id')
        client_ids_param = request.args.get('client_ids')  # 逗号分隔的client_id列表

        with get_db_readonly_safe() as conn:
            cursor = conn.cursor()

            # 构建查询
            if client_ids_param:
                client_ids = [cid.strip() for cid in client_ids_param.split(',') if cid.strip()]
                placeholders = ','.join(['%s' for _ in client_ids])
                cursor.execute(f'''
                    SELECT c.client_id, c.hostname, c.local_ip, g.name as group_name,
                           c.last_report, c.created_at
                    FROM clients c
                    LEFT JOIN `groups` g ON c.group_id = g.id
                    WHERE c.client_id IN ({placeholders})
                    ORDER BY c.last_report DESC
                ''', client_ids)
            elif group_id == 'ungrouped':
                cursor.execute('''
                    SELECT c.client_id, c.hostname, c.local_ip, g.name as group_name,
                           c.last_report, c.created_at
                    FROM clients c
                    LEFT JOIN `groups` g ON c.group_id = g.id
                    WHERE c.group_id IS NULL
                    ORDER BY c.last_report DESC
                ''')
            elif group_id:
                cursor.execute('''
                    SELECT c.client_id, c.hostname, c.local_ip, g.name as group_name,
                           c.last_report, c.created_at
                    FROM clients c
                    LEFT JOIN `groups` g ON c.group_id = g.id
                    WHERE c.group_id = %s
                    ORDER BY c.last_report DESC
                ''', (group_id,))
            else:
                cursor.execute('''
                    SELECT c.client_id, c.hostname, c.local_ip, g.name as group_name,
                           c.last_report, c.created_at
                    FROM clients c
                    LEFT JOIN `groups` g ON c.group_id = g.id
                    ORDER BY c.last_report DESC
                ''')

            clients = cursor.fetchall()

            # 创建Excel工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = '客户端列表'

            # 表头样式
            header_font = Font(bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='667eea', end_color='667eea', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center')

            # 表头
            headers = ['主机名', '客户端ID', 'IP地址', '分组', '最后上报时间', '创建时间',
                       'CPU', '内存', '硬盘', '显卡', '运行时间', '温度(最高)', '风扇(最高)', '电压']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            # 填充数据
            for row_idx, client in enumerate(clients, 2):
                client_dict = dict(client)
                client_id = client_dict['client_id']

                # 获取最新硬件信息
                cursor.execute('''
                    SELECT report_data FROM hardware_reports
                    WHERE client_id = %s ORDER BY timestamp DESC LIMIT 1
                ''', (client_id,))
                report = cursor.fetchone()

                # 提取关键指标
                cpu_str = '-'
                mem_str = '-'
                disk_str = '-'
                gpu_str = '-'
                uptime_str = '-'
                temp_str = '-'
                fan_str = '-'
                voltage_str = '-'

                if report:
                    hardware = json.loads(report['report_data'])

                    # CPU
                    if hardware.get('cpu') and isinstance(hardware['cpu'], list):
                        cpu_names = [c.get('name', '?') for c in hardware['cpu']]
                        cpu_cores = [f"{c.get('cores', '?')}核" for c in hardware['cpu']]
                        cpu_str = ' | '.join([f"{n}({cores})" for n, cores in zip(cpu_names, cpu_cores)])

                    # 内存
                    if hardware.get('memory'):
                        total = hardware['memory'].get('total_capacity', 0)
                        if total:
                            mem_str = f"{total / (1024**3):.1f} GB"
                        elif hardware['memory'].get('modules'):
                            total = sum(m.get('capacity', 0) for m in hardware['memory']['modules'])
                            mem_str = f"{total / (1024**3):.1f} GB"

                    # 硬盘
                    if hardware.get('disk') and isinstance(hardware['disk'], list):
                        disk_models = [d.get('model', '?') for d in hardware['disk']]
                        disk_sizes = [f"{d.get('size', 0) / (1024**3):.0f}GB" for d in hardware['disk']]
                        disk_str = ' | '.join([f"{m}({s})" for m, s in zip(disk_models, disk_sizes)])

                    # 显卡
                    if hardware.get('gpu') and isinstance(hardware['gpu'], list):
                        gpu_names = [g.get('name', '?') for g in hardware['gpu']]
                        gpu_str = ' | '.join(gpu_names)

                    # 运行时间
                    if hardware.get('uptime') and not hardware['uptime'].get('error'):
                        uptime_str = hardware['uptime'].get('uptime_human', '-')

                    # 温度（取最高值）
                    if hardware.get('temperature') and hardware['temperature'].get('sensors'):
                        temps = hardware['temperature']['sensors']
                        if temps:
                            max_temp = max(s.get('value', 0) for s in temps)
                            temp_str = f"{max_temp}°C"

                    # 风扇（取最高转速）
                    if hardware.get('fan') and hardware['fan'].get('sensors'):
                        fans = hardware['fan']['sensors']
                        if fans:
                            max_fan = max(s.get('value', 0) for s in fans)
                            fan_str = f"{max_fan} RPM"

                    # 电压
                    if hardware.get('voltage') and hardware['voltage'].get('sensors'):
                        voltages = hardware['voltage']['sensors']
                        if voltages:
                            voltage_str = ' | '.join([f"{s.get('name','?')}:{s.get('value',0)}V" for s in voltages[:3]])

                ws.cell(row=row_idx, column=1, value=client_dict.get('hostname') or client_id)
                ws.cell(row=row_idx, column=2, value=client_id)
                ws.cell(row=row_idx, column=3, value=client_dict.get('local_ip') or '-')
                ws.cell(row=row_idx, column=4, value=client_dict.get('group_name') or '未分组')
                ws.cell(row=row_idx, column=5, value=client_dict.get('last_report') or '-')
                ws.cell(row=row_idx, column=6, value=client_dict.get('created_at') or '-')
                ws.cell(row=row_idx, column=7, value=cpu_str)
                ws.cell(row=row_idx, column=8, value=mem_str)
                ws.cell(row=row_idx, column=9, value=disk_str)
                ws.cell(row=row_idx, column=10, value=gpu_str)
                ws.cell(row=row_idx, column=11, value=uptime_str)
                ws.cell(row=row_idx, column=12, value=temp_str)
                ws.cell(row=row_idx, column=13, value=fan_str)
                ws.cell(row=row_idx, column=14, value=voltage_str)

            # 调整列宽
            col_widths = [15, 20, 16, 15, 22, 22, 40, 12, 40, 30, 20, 15, 15, 30]
            for i, width in enumerate(col_widths, 1):
                ws.column_dimensions[chr(64 + i) if i <= 26 else 'A' + chr(64 + i - 26)].width = width

        # 保存到内存并返回
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'hardware_report_{datetime.now(TZ_CST).strftime("%Y%m%d_%H%M%S")}.xlsx'
        )

    except Exception as e:
        logger.error(f'导出Excel失败: {e}')
        return jsonify({'error': str(e)}), 500


@app.route('/api/clients/batch-group', methods=['PUT'])
@login_required
def batch_assign_group():
    """批量分配客户端到分组"""
    try:
        data = request.json
        client_ids = data.get('client_ids', [])
        group_id = data.get('group_id')

        # 输入验证
        if not client_ids:
            return jsonify({'error': '请选择要操作的客户端'}), 400

        # 限制批量操作数量，防止SQL过长
        MAX_BATCH_SIZE = 1000
        if len(client_ids) > MAX_BATCH_SIZE:
            return jsonify({'error': f'批量操作数量不能超过{MAX_BATCH_SIZE}个'}), 400

        # 验证client_ids格式（只允许字母、数字、下划线、连字符）
        import re
        for cid in client_ids:
            if not re.match(r'^[a-zA-Z0-9_-]+$', str(cid)):
                return jsonify({'error': f'无效的客户端ID: {cid}'}), 400

        with get_db_safe() as conn:
            cursor = conn.cursor()

            placeholders = ','.join(['%s' for _ in client_ids])
            cursor.execute(f'''
                UPDATE clients SET group_id = %s
                WHERE client_id IN ({placeholders})
            ''', [group_id] + client_ids)

            affected = cursor.rowcount
            conn.commit()

        return jsonify({
            'status': 'success',
            'affected': affected,
            'message': f'成功将 {affected} 个客户端分配到分组'
        })

    except Exception as e:
        logger.error(f'批量分配分组失败: {e}')
        return jsonify({'error': str(e)}), 500


@app.route('/api/client/<client_id>/history', methods=['GET'])
@login_required
def get_client_history(client_id):
    """获取客户端硬件采集历史记录（最近10条）"""
    try:
        with get_db_safe() as conn:
            cursor = conn.cursor()

            # 验证客户端存在
            cursor.execute('SELECT client_id FROM clients WHERE client_id = %s', (client_id,))
            if not cursor.fetchone():
                return jsonify({'error': '客户端不存在'}), 404

            # 获取最近10条历史记录
            cursor.execute('''
                SELECT cpu_info, memory_info, disk_info, gpu_info, snapshot, timestamp
                FROM hardware_history
                WHERE client_id = %s
                ORDER BY timestamp DESC
                LIMIT 10
            ''', (client_id,))

            history = []
            for row in cursor.fetchall():
                row_dict = dict(row)
                # 解析快照
                if row_dict.get('snapshot'):
                    snapshot = json.loads(row_dict['snapshot'])
                else:
                    snapshot = {}
                row_dict['snapshot'] = snapshot
                history.append(row_dict)

        return jsonify({'status': 'success', 'data': history})

    except Exception as e:
        logger.error(f'获取客户端历史失败: {e}')
        return jsonify({'error': str(e)}), 500


@app.route('/api/clients/batch-delete', methods=['DELETE'])
@login_required
def batch_delete_clients():
    """批量删除客户端"""
    try:
        data = request.json
        client_ids = data.get('client_ids', [])

        if not client_ids:
            return jsonify({'error': '请选择要删除的客户端'}), 400

        with get_db_safe() as conn:
            cursor = conn.cursor()

            placeholders = ','.join(['%s' for _ in client_ids])
            cursor.execute(f'DELETE FROM clients WHERE client_id IN ({placeholders})', client_ids)
            affected = cursor.rowcount
            conn.commit()

        return jsonify({
            'status': 'success',
            'affected': affected,
            'message': f'成功删除 {affected} 个客户端'
        })

    except Exception as e:
        logger.error(f'批量删除客户端失败: {e}')
        return jsonify({'error': str(e)}), 500


@app.route('/api/client/<client_id>/baseline', methods=['GET'])
@login_required
def get_client_baseline(client_id):
    """获取客户端的硬件基准信息（含手动基准配置与条目 key）"""
    try:
        from compare_engine import _item_key

        with get_db_safe() as conn:
            cursor = conn.cursor()

            cursor.execute('SELECT * FROM client_baselines WHERE client_id = %s', (client_id,))
            baseline = cursor.fetchone()

            if not baseline:
                return jsonify({'status': 'not_found', 'message': '该客户端尚未建立基准'})

            baseline_dict = dict(baseline)
            # 解析JSON字段
            for key in ['cpu_snapshot', 'gpu_snapshot', 'memory_snapshot', 'disk_snapshot']:
                if baseline_dict.get(key):
                    baseline_dict[key] = json.loads(baseline_dict[key])
            # 解析手动基准字段
            baseline_dict['selected_items'] = _parse_json_field(baseline_dict.get('selected_items')) or {}
            baseline_dict['manual_overrides'] = _parse_json_field(baseline_dict.get('manual_overrides')) or {}
            baseline_dict['baseline_mode'] = baseline_dict.get('baseline_mode') or 'auto'

        # 为前端构建「条目 + item_key」列表，供勾选/改值使用
        items = {}
        try:
            cpu_items = baseline_dict.get('cpu_snapshot') or []
            items['cpu'] = [{'key': _item_key(i, 'cpu'), 'data': i} for i in cpu_items if isinstance(i, dict)]

            gpu_items = baseline_dict.get('gpu_snapshot') or []
            items['gpu'] = [{'key': _item_key(i, 'gpu'), 'data': i} for i in gpu_items if isinstance(i, dict)]

            disk_items = baseline_dict.get('disk_snapshot') or []
            items['disk'] = [{'key': _item_key(i, 'disk'), 'data': i} for i in disk_items if isinstance(i, dict)]

            mem_snap = baseline_dict.get('memory_snapshot') or {}
            mem_modules = mem_snap.get('modules', []) if isinstance(mem_snap, dict) else []
            items['memory'] = [{'key': _item_key(m, 'memory'), 'data': m} for m in mem_modules if isinstance(m, dict)]
        except Exception as e:
            logger.warning(f'构建基准条目 key 失败 {client_id}: {e}')
            items = {}

        baseline_dict['items'] = items

        return jsonify({'status': 'success', 'data': baseline_dict})

    except Exception as e:
        logger.error(f'获取客户端基准失败: {e}')
        return jsonify({'error': str(e)}), 500


@app.route('/api/client/<client_id>/baseline', methods=['POST'])
@login_required
def set_client_baseline(client_id):
    """手动设置/重置客户端的硬件基准（使用当前最新上报数据，可选手动勾选+改值）

    body 可选字段：
        selected_items    dict {hw_type: [item_key, ...]}  勾选纳入对比的条目（空=全量）
        manual_overrides  dict {hw_type: {item_key: {field: value}}}  手动覆盖字段值
        baseline_mode     'auto' | 'manual'
        cpu_snapshot / gpu_snapshot / memory_snapshot / disk_snapshot
                          可选：直接复用已有基准快照（保证勾选/覆盖的 item_key 与快照一致）。
                          缺省时用「最新上报数据」重建快照。
    """
    try:
        data = request.json or {}
        selected_items = data.get('selected_items')
        manual_overrides = data.get('manual_overrides')
        baseline_mode = data.get('baseline_mode', 'auto')

        # 规范化：手动模式需要 mode 标记；若提供了勾选/覆盖则自动视为 manual
        if selected_items or manual_overrides:
            baseline_mode = 'manual'
        if baseline_mode not in ('auto', 'manual'):
            baseline_mode = 'auto'

        with get_db_safe() as conn:
            cursor = conn.cursor()

            # 验证客户端存在
            cursor.execute('SELECT client_id FROM clients WHERE client_id = %s', (client_id,))
            if not cursor.fetchone():
                return jsonify({'error': '客户端不存在'}), 404

            # 若调用方显式传入了快照（来自 GET /baseline），直接复用，保证 item_key 一致
            snapshot_keys = ['cpu_snapshot', 'gpu_snapshot', 'memory_snapshot', 'disk_snapshot']
            if all(k in data and data[k] is not None for k in snapshot_keys):
                cpu_info = json.dumps(data['cpu_snapshot'], ensure_ascii=False)
                gpu_info = json.dumps(data['gpu_snapshot'], ensure_ascii=False)
                mem_info = json.dumps(data['memory_snapshot'], ensure_ascii=False)
                disk_info = json.dumps(data['disk_snapshot'], ensure_ascii=False)
            else:
                # 获取最新的硬件报告
                cursor.execute('''
                    SELECT report_data FROM hardware_reports
                    WHERE client_id = %s ORDER BY timestamp DESC LIMIT 1
                ''', (client_id,))
                report = cursor.fetchone()

                if not report:
                    return jsonify({'error': '该客户端尚未上报任何硬件数据'}), 400

                hardware_info = json.loads(report['report_data'])

                # 提取关键指标
                cpu_info = json.dumps(hardware_info.get('cpu', []), ensure_ascii=False) if hardware_info.get('cpu') else ''
                mem_info = json.dumps(hardware_info.get('memory', {}), ensure_ascii=False) if hardware_info.get('memory') else ''
                disk_info = json.dumps(hardware_info.get('disk', []), ensure_ascii=False) if hardware_info.get('disk') else ''
                gpu_info = json.dumps(hardware_info.get('gpu', []), ensure_ascii=False) if hardware_info.get('gpu') else ''

            sel_json = json.dumps(selected_items, ensure_ascii=False) if selected_items else None
            ovr_json = json.dumps(manual_overrides, ensure_ascii=False) if manual_overrides else None

            # 插入或更新基准
            cursor.execute('''
                INSERT INTO client_baselines
                    (client_id, cpu_snapshot, gpu_snapshot, memory_snapshot, disk_snapshot,
                     selected_items, manual_overrides, baseline_mode)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                ON DUPLICATE KEY UPDATE
                    cpu_snapshot = VALUES(cpu_snapshot),
                    gpu_snapshot = VALUES(gpu_snapshot),
                    memory_snapshot = VALUES(memory_snapshot),
                    disk_snapshot = VALUES(disk_snapshot),
                    selected_items = VALUES(selected_items),
                    manual_overrides = VALUES(manual_overrides),
                    baseline_mode = VALUES(baseline_mode),
                    baseline_timestamp = CURRENT_TIMESTAMP
            ''', (client_id, cpu_info, gpu_info, mem_info, disk_info, sel_json, ovr_json, baseline_mode))

            conn.commit()

        return jsonify({'status': 'success', 'message': '基准已更新', 'baseline_mode': baseline_mode})

    except Exception as e:
        logger.error(f'设置客户端基准失败: {e}')
        return jsonify({'error': str(e)}), 500


@app.route('/api/client/<client_id>/baseline/sync', methods=['POST'])
@login_required
def sync_client_baseline(client_id):
    """把指定客户端的基准（含手动配置）同步到其他设备。

    body:
        target_client_ids  list  手动勾选的目标客户端 ID 列表
        group_id           int   同步到整个分组（与 target_client_ids 二选一或可同时）
    """
    try:
        data = request.json or {}
        target_client_ids = data.get('target_client_ids') or []
        group_id = data.get('group_id')

        with get_db_safe() as conn:
            cursor = conn.cursor()

            # 读取源基准
            cursor.execute('SELECT * FROM client_baselines WHERE client_id = %s', (client_id,))
            src = cursor.fetchone()
            if not src:
                return jsonify({'error': '源客户端尚未建立基准'}), 400

            # 汇总目标客户端 ID
            targets = set(target_client_ids)
            if group_id:
                cursor.execute('SELECT client_id FROM clients WHERE group_id = %s', (group_id,))
                for row in cursor.fetchall():
                    targets.add(row['client_id'])
            targets.discard(client_id)  # 排除自身

            if not targets:
                return jsonify({'error': '未指定任何目标设备'}), 400

            src_dict = dict(src)
            target_list = sorted(targets)
            synced = len(target_list)

            # 单条批量 upsert：多行 VALUES，一次往返完成（替代逐台循环，大幅提速）
            values = []
            params = []
            for tid in target_list:
                values.append('(%s, %s, %s, %s, %s, %s, %s, %s)')
                params.extend([
                    tid,
                    src_dict.get('cpu_snapshot'),
                    src_dict.get('gpu_snapshot'),
                    src_dict.get('memory_snapshot'),
                    src_dict.get('disk_snapshot'),
                    src_dict.get('selected_items'),
                    src_dict.get('manual_overrides'),
                    src_dict.get('baseline_mode') or 'auto',
                ])

            cursor.execute(f'''
                INSERT INTO client_baselines
                    (client_id, cpu_snapshot, gpu_snapshot, memory_snapshot, disk_snapshot,
                     selected_items, manual_overrides, baseline_mode)
                VALUES {','.join(values)}
                ON DUPLICATE KEY UPDATE
                    cpu_snapshot = VALUES(cpu_snapshot),
                    gpu_snapshot = VALUES(gpu_snapshot),
                    memory_snapshot = VALUES(memory_snapshot),
                    disk_snapshot = VALUES(disk_snapshot),
                    selected_items = VALUES(selected_items),
                    manual_overrides = VALUES(manual_overrides),
                    baseline_mode = VALUES(baseline_mode),
                    baseline_timestamp = CURRENT_TIMESTAMP
            ''', params)

            conn.commit()

        return jsonify({'status': 'success', 'message': f'已同步到 {synced} 台设备', 'synced': synced})

    except Exception as e:
        logger.error(f'同步基准失败: {e}')
        return jsonify({'error': str(e)}), 500


@app.route('/api/client/<client_id>/alerts', methods=['GET'])
@login_required
def get_client_alerts(client_id):
    """获取指定客户端的告警记录"""
    try:
        resolved = request.args.get('resolved')  # 'true' or 'false' or None for all

        with get_db_safe() as conn:
            cursor = conn.cursor()

            query = '''
                SELECT a.*, c.hostname, c.local_ip
                FROM alert_records a
                LEFT JOIN clients c ON a.client_id = c.client_id
                WHERE a.client_id = %s
            '''
            params = [client_id]

            if resolved is not None:
                query += ' AND a.resolved = %s'
                params.append(1 if resolved == 'true' else 0)

            query += ' ORDER BY a.created_at DESC'

            cursor.execute(query, params)
            alerts = [dict(row) for row in cursor.fetchall()]

            # 解析alert_detail JSON
            for alert in alerts:
                if alert.get('alert_detail'):
                    alert['alert_detail'] = json.loads(alert['alert_detail'])

        return jsonify({'status': 'success', 'data': alerts})

    except Exception as e:
        logger.error(f'获取客户端告警失败: {e}')
        return jsonify({'error': str(e)}), 500


@app.route('/api/alerts', methods=['GET'])
@login_required
def get_all_alerts():
    """获取所有告警记录（支持分页、过滤和关键词搜索）"""
    try:
        page = int(request.args.get('page', 1))
        per_page = int(request.args.get('per_page', 20))
        resolved = request.args.get('resolved')
        # 关键词搜索（主机名/client_id/IP/告警详情），与客户端列表搜索同策略：
        # 带搜索时不走缓存，避免关键词组合污染缓存
        search = (request.args.get('search') or '').strip()

        if not search:
            # 短缓存（5s）：告警列表是高频读取热路径，命中直接返回，绕过 DB 查询
            cache_key = f'hwmon:alerts:{page}:{per_page}:{resolved or "all"}'
            cached = get_cache(cache_key)
            if cached:
                return jsonify(cached)
        else:
            cache_key = None

        with get_db_safe() as conn:
            cursor = conn.cursor()

            # 检查表是否存在
            try:
                cursor.execute("SELECT 1 FROM alert_records LIMIT 1")
            except Exception:
                conn.rollback()
                try:
                    init_tables()
                except Exception:
                    pass
                return jsonify({'status': 'success', 'data': [], 'pagination': {'page': 1, 'per_page': per_page, 'total': 0, 'pages': 0}})

            query = '''
                SELECT a.*, c.hostname, c.local_ip
                FROM alert_records a
                LEFT JOIN clients c ON a.client_id = c.client_id
                WHERE 1=1
            '''
            params = []

            if resolved is not None:
                query += ' AND a.resolved = %s'
                params.append(1 if resolved == 'true' else 0)

            # 关键词模糊搜索：主机名 / client_id / IP / 告警详情
            if search:
                like = f'%{search}%'
                query += ''' AND (a.client_id LIKE %s OR c.hostname LIKE %s
                            OR c.local_ip LIKE %s OR a.alert_detail LIKE %s
                            OR a.alert_type LIKE %s)'''
                params.extend([like, like, like, like, like])

            # 获取总数
            count_query = query.replace('SELECT a.*, c.hostname, c.local_ip', 'SELECT COUNT(*) as total')
            cursor.execute(count_query, params)
            total = cursor.fetchone()['total']

            # 准确统计「未解决告警」总数（供前端告警中心数字使用，
            # 避免 per_page 截断导致列表长度不准）
            cursor.execute('SELECT COUNT(*) as cnt FROM alert_records WHERE resolved = 0')
            unresolved_total = cursor.fetchone()['cnt']

            # 分页查询
            query += ' ORDER BY a.created_at DESC LIMIT %s OFFSET %s'
            params.extend([per_page, (page - 1) * per_page])

            cursor.execute(query, params)
            alerts = [dict(row) for row in cursor.fetchall()]

            # 解析alert_detail JSON
            for alert in alerts:
                if alert.get('alert_detail'):
                    alert['alert_detail'] = json.loads(alert['alert_detail'])

        response_data = {
            'status': 'success',
            'data': alerts,
            'unresolved_total': unresolved_total,
            'pagination': {
                'page': page,
                'per_page': per_page,
                'total': total,
                'pages': (total + per_page - 1) // per_page
            }
        }
        # 带搜索时 cache_key 为 None，不写缓存
        if cache_key:
            set_cache(cache_key, response_data, ttl=5)
        return jsonify(response_data)

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/alerts/<int:alert_id>', methods=['PUT'])
@login_required
def resolve_alert(alert_id):
    """标记单个告警为已解决"""
    try:
        with get_db_safe() as conn:
            cursor = conn.cursor()

            cursor.execute('UPDATE alert_records SET resolved = 1 WHERE id = %s', (alert_id,))

            if cursor.rowcount == 0:
                return jsonify({'error': '告警记录不存在'}), 404

            conn.commit()

        return jsonify({'status': 'success', 'message': '告警已标记为已解决'})

    except Exception as e:
        logger.error(f'标记告警失败: {e}')
        return jsonify({'error': str(e)}), 500


@app.route('/api/alerts/<int:alert_id>', methods=['DELETE'])
@login_required
def delete_alert(alert_id):
    """删除单个告警记录"""
    try:
        with get_db_safe() as conn:
            cursor = conn.cursor()
            cursor.execute('DELETE FROM alert_records WHERE id = %s', (alert_id,))

            if cursor.rowcount == 0:
                return jsonify({'error': '告警记录不存在'}), 404

            conn.commit()

        return jsonify({'status': 'success', 'message': '告警已删除'})

    except Exception as e:
        logger.error(f'删除告警失败: {e}')
        return jsonify({'error': str(e)}), 500


@app.route('/api/alerts/batch-resolve', methods=['PUT'])
@login_required
def batch_resolve_alerts():
    """批量标记告警为已解决"""
    try:
        data = request.json
        alert_ids = data.get('alert_ids', [])

        if not alert_ids:
            return jsonify({'error': '请选择要解决的告警'}), 400

        with get_db_safe() as conn:
            cursor = conn.cursor()

            placeholders = ','.join(['%s' for _ in alert_ids])
            cursor.execute(f'UPDATE alert_records SET resolved = 1 WHERE id IN ({placeholders})', alert_ids)
            affected = cursor.rowcount

            conn.commit()

        return jsonify({
            'status': 'success',
            'affected': affected,
            'message': f'成功标记 {affected} 个告警为已解决'
        })

    except Exception as e:
        logger.error(f'批量标记告警失败: {e}')
        return jsonify({'error': str(e)}), 500


@app.route('/api/alerts/resolve-all', methods=['PUT'])
@login_required
def resolve_all_alerts():
    """一键标记所有未解决告警为已解决"""
    try:
        with get_db_safe() as conn:
            cursor = conn.cursor()
            cursor.execute('UPDATE alert_records SET resolved = 1 WHERE resolved = 0')
            affected = cursor.rowcount
            conn.commit()

        return jsonify({
            'status': 'success',
            'affected': affected,
            'message': f'已解决 {affected} 条未解决告警'
        })

    except Exception as e:
        logger.error(f'一键解决告警失败: {e}')
        return jsonify({'error': str(e)}), 500


@app.route('/api/email-config', methods=['GET'])
@login_required
def get_email_config_api():
    """获取邮件配置"""
    try:
        with get_db_readonly_safe() as conn:
            config = get_email_config(conn)

        if not config:
            return jsonify({'error': '邮件配置不存在'}), 404

        # 隐藏密码字段（返回时不显示完整密码）
        config_copy = dict(config)
        if config_copy.get('smtp_password'):
            config_copy['smtp_password'] = '******'

        return jsonify({'status': 'success', 'data': config_copy})

    except Exception as e:
        logger.error(f'获取邮件配置失败: {e}')
        return jsonify({'error': str(e)}), 500


@app.route('/api/email-config', methods=['PUT'])
@login_required
def update_email_config():
    """更新邮件配置"""
    try:
        data = request.json

        with get_db_safe() as conn:
            cursor = conn.cursor()

            # 检查是否需要保留原密码
            if data.get('smtp_password') == '******':
                # 获取原密码
                cursor.execute('SELECT smtp_password FROM email_config WHERE id = 1')
                row = cursor.fetchone()
                old_password = row['smtp_password'] if row else ''
                data['smtp_password'] = old_password

            cursor.execute('''
                UPDATE email_config SET
                    smtp_host = %s,
                    smtp_port = %s,
                    smtp_user = %s,
                    smtp_password = %s,
                    sender_name = %s,
                    recipients = %s,
                    enabled = %s
                WHERE id = 1
            ''', (
                data.get('smtp_host', 'smtp.qq.com'),
                data.get('smtp_port', 465),
                data.get('smtp_user', ''),
                data.get('smtp_password', ''),
                data.get('sender_name', '硬件监控系统'),
                json.dumps(data.get('recipients', [])),
                1 if data.get('enabled') else 0
            ))

            conn.commit()

        return jsonify({'status': 'success', 'message': '邮件配置已更新'})

    except Exception as e:
        logger.error(f'更新邮件配置失败: {e}')
        return jsonify({'error': str(e)}), 500


@app.route('/api/email-config/test', methods=['POST'])
@login_required
def test_email_config():
    """测试邮件配置（发送测试邮件）"""
    try:
        data = request.json

        import smtplib
        from email.mime.text import MIMEText
        from email.mime.multipart import MIMEMultipart

        smtp_host = data.get('smtp_host', 'smtp.qq.com')
        smtp_port = data.get('smtp_port', 465)
        smtp_user = data.get('smtp_user', '')
        smtp_password = data.get('smtp_password', '')
        sender_name = data.get('sender_name', '硬件监控系统')
        test_recipient = data.get('test_recipient', '')

        if not smtp_user or not smtp_password or not test_recipient:
            return jsonify({'error': '请填写完整的SMTP配置和测试收件人'}), 400

        # 构建测试邮件
        html_body = f'''
        <html><body style="font-family:Microsoft YaHei,Arial,sans-serif;">
        <h2 style="color:#38a169;">【邮件配置测试】</h2>
        <p>这是一封来自硬件监控系统的测试邮件。</p>
        <p>如果您收到此邮件，说明SMTP配置正确。</p>
        <p style="color:#718096; margin-top:20px;">发送时间: {datetime.now(TZ_CST).strftime("%Y-%m-%d %H:%M:%S")}</p>
        </body></html>
        '''

        msg = MIMEMultipart('alternative')
        msg['Subject'] = '【测试邮件】硬件监控系统配置测试'
        msg['From'] = f'{sender_name} <{smtp_user}>'
        msg['To'] = test_recipient
        msg.attach(MIMEText(html_body, 'html', 'utf-8'))

        # 发送邮件
        server = smtplib.SMTP_SSL(smtp_host, smtp_port)
        server.login(smtp_user, smtp_password)
        server.sendmail(smtp_user, [test_recipient], msg.as_string())
        server.quit()

        return jsonify({'status': 'success', 'message': '测试邮件发送成功'})

    except smtplib.SMTPAuthenticationError:
        return jsonify({'error': 'SMTP认证失败，请检查用户名和密码'}), 400
    except smtplib.SMTPConnectError:
        return jsonify({'error': '无法连接到SMTP服务器，请检查主机和端口'}), 400
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'邮件发送失败: {str(e)}'}), 500


@app.route('/api/alert-settings', methods=['GET'])
@login_required
def get_alert_settings():
    """获取告警设置（带缓存）"""
    try:
        # 尝试从缓存获取
        cache_key = 'hwmon:alert_settings:1'
        cached = get_cache(cache_key)
        if cached:
            return jsonify({'status': 'success', 'data': cached})

        with get_db_safe() as conn:
            cursor = conn.cursor()
            cursor.execute('SELECT * FROM alert_settings WHERE id = 1')
            row = cursor.fetchone()

        if row:
            settings = dict(row)
        else:
            settings = {
                'alert_enabled': 1,
                'email_enabled': 0,
                'monitor_cpu': 1,
                'monitor_gpu': 1,
                'monitor_memory': 1,
                'monitor_disk': 1,
                'monitor_network': 0,
                'monitor_motherboard': 0,
                'monitor_bios': 0,
                'monitor_temperature': 0,
                'monitor_fan': 0,
                'monitor_voltage': 0
            }

        # 写入缓存（5分钟TTL）
        set_cache(cache_key, settings, ttl=300)

        return jsonify({'status': 'success', 'data': settings})

    except Exception as e:
        logger.error(f'获取告警设置失败: {e}')
        return jsonify({'error': str(e)}), 500


@app.route('/api/alert-settings', methods=['PUT'])
@login_required
def update_alert_settings():
    """更新告警设置"""
    try:
        data = request.json

        with get_db_safe() as conn:
            cursor = conn.cursor()

            cursor.execute('''
                UPDATE alert_settings SET
                    alert_enabled = %s,
                    email_enabled = %s,
                    monitor_cpu = %s,
                    monitor_gpu = %s,
                    monitor_memory = %s,
                    monitor_disk = %s,
                    monitor_network = %s,
                    monitor_motherboard = %s,
                    monitor_bios = %s,
                    monitor_temperature = %s,
                    monitor_fan = %s,
                    monitor_voltage = %s,
                    updated_at = CURRENT_TIMESTAMP
                WHERE id = 1
            ''', (
                1 if data.get('alert_enabled') else 0,
                1 if data.get('email_enabled') else 0,
                1 if data.get('monitor_cpu') else 0,
                1 if data.get('monitor_gpu') else 0,
                1 if data.get('monitor_memory') else 0,
                1 if data.get('monitor_disk') else 0,
                1 if data.get('monitor_network') else 0,
                1 if data.get('monitor_motherboard') else 0,
                1 if data.get('monitor_bios') else 0,
                1 if data.get('monitor_temperature') else 0,
                1 if data.get('monitor_fan') else 0,
                1 if data.get('monitor_voltage') else 0
            ))

            conn.commit()

        # 清除缓存
        on_alert_settings_update()

        return jsonify({'status': 'success', 'message': '告警设置已更新'})

    except Exception as e:
        logger.error(f'更新告警设置失败: {e}')
        return jsonify({'error': str(e)}), 500


@app.route('/api/config', methods=['GET'])
@login_required
def get_collect_config():
    """获取采集配置"""
    return jsonify({
        'status': 'success',
        'data': {
            'max_workers': COLLECT_CONFIG['max_workers'],
            'timeout': COLLECT_CONFIG['timeout'],
            'retry_times': COLLECT_CONFIG['retry_times']
        }
    })


# =================================================================
# 仪表盘统计 API
# =================================================================

@app.route('/api/dashboard', methods=['GET'])
@login_required
def get_dashboard():
    """获取仪表盘统计数据（带缓存）"""
    try:
        # 尝试从缓存获取
        cache_key = 'hwmon:dashboard:stats'
        cached = get_cache(cache_key)
        if cached:
            return jsonify(cached)

        with get_db_readonly_safe() as conn:
            cursor = conn.cursor()

            # 合并查询：一次获取所有计数（减少数据库往返）
            cursor.execute('''
                SELECT
                    (SELECT COUNT(*) FROM clients) as total_clients,
                    (SELECT COUNT(*) FROM clients WHERE last_report >= DATE_SUB(NOW(), INTERVAL 24 HOUR)) as online_clients,
                    (SELECT COUNT(*) FROM `groups`) as total_groups,
                    (SELECT COUNT(*) FROM alert_records WHERE resolved = 0) as unresolved_alerts,
                    (SELECT COUNT(*) FROM process_alert_records WHERE resolved = 0) as process_alerts,
                    (SELECT COUNT(*) FROM probe_alerts WHERE resolved = 0) as probe_alerts,
                    (SELECT COUNT(*) FROM process_alert_records WHERE ai_analyzed = 1) as ai_analyzed
            ''')
            counts = cursor.fetchone()

            # 最近 10 条上报记录
            cursor.execute('''
                SELECT c.client_id, c.hostname, c.local_ip, c.last_report, g.name as group_name
                FROM clients c
                LEFT JOIN `groups` g ON c.group_id = g.id
                ORDER BY c.last_report DESC
                LIMIT 10
            ''')
            recent_reports = [dict(row) for row in cursor.fetchall()]

            # 分组统计
            cursor.execute('''
                SELECT g.name, g.id, COUNT(c.id) as client_count
                FROM `groups` g
                LEFT JOIN clients c ON g.id = c.group_id
                GROUP BY g.id
                ORDER BY g.name
            ''')
            group_stats = [dict(row) for row in cursor.fetchall()]

        response_data = {
            'status': 'success',
            'data': {
                'total_clients': counts['total_clients'],
                'online_clients': counts['online_clients'],
                'offline_clients': counts['total_clients'] - counts['online_clients'],
                'total_groups': counts['total_groups'],
                'unresolved_alerts': counts['unresolved_alerts'],
                'process_alerts': counts['process_alerts'],
                'ai_analyzed': counts['ai_analyzed'],
                'probe_alerts': counts['probe_alerts'],
                'recent_reports': recent_reports,
                'group_stats': group_stats,
            }
        }

        # 写入缓存（5秒TTL：让侧栏告警角标更实时）
        set_cache(cache_key, response_data, ttl=5)

        return jsonify(response_data)
    except Exception as e:
        logger.error(f'获取仪表盘数据失败: {e}')
        return jsonify({'error': str(e)}), 500


# =================================================================
# 进程告警 API
# =================================================================

def _async_ai_analyze(alert_id, client_id, alert_data):
    """异步 AI 分析（在线程池中执行，不阻塞主流程）

    修复说明：此前该函数从未定义，导致 /api/process-alert 提交任务时
    抛出 NameError（异常被线程池吞掉），AI 自动研判完全失效。
    现参照 tasks.py 中的 async_ai_analyze 逻辑补齐。
    """
    try:
        conn = get_db()
        cursor = conn.cursor()

        # 检查是否在 AI 监控列表
        cursor.execute('SELECT id FROM ai_monitored_hosts WHERE client_id = %s AND enabled = 1', (client_id,))
        if not cursor.fetchone():
            conn.close()
            return

        # 获取 AI 配置
        cursor.execute('SELECT * FROM ai_config WHERE id = 1')
        ai_cfg = cursor.fetchone()
        if not ai_cfg or not ai_cfg.get('enabled') or not ai_cfg.get('auto_analyze'):
            conn.close()
            return

        alert_json = json.loads(alert_data) if isinstance(alert_data, str) else alert_data
        result = analyze_process_alert(alert_json, dict(ai_cfg))

        if result:
            cursor.execute('''
                UPDATE process_alert_records
                SET ai_analyzed = 1, ai_result = %s
                WHERE id = %s
            ''', (json.dumps(result, ensure_ascii=False), alert_id))
            conn.commit()

        conn.close()

    except Exception as e:
        logger.warning(f'AI 异步分析失败 (alert_id={alert_id}): {e}')


@app.route('/api/process-alert', methods=['POST'])
def receive_process_alert():
    """接收客户端进程告警（同步处理）"""
    try:
        data = request.json
        client_id = data.get('client_id')
        hostname = data.get('hostname', '')
        local_ip = data.get('local_ip', '')
        # APIPA/无效地址兜底：169.254.x.x（DHCP 失败自动地址）、回环等改用真实来源 IP
        if not local_ip or local_ip.startswith('169.254.') or local_ip in ('127.0.0.1', '0.0.0.0', '::1'):
            local_ip = request.remote_addr or ''
        alerts = data.get('alerts', [])
        system_summary = data.get('system_summary', {})

        if not client_id:
            return jsonify({'error': '缺少 client_id'}), 400

        if not alerts:
            return jsonify({'error': '告警列表为空'}), 400

        conn = get_db()
        cursor = conn.cursor()

        # 确保客户端存在
        current_time = datetime.now(TZ_CST).strftime('%Y-%m-%d %H:%M:%S')
        cursor.execute('''
            INSERT INTO clients (client_id, hostname, local_ip, last_report)
            VALUES (%s, %s, %s, %s)
            ON DUPLICATE KEY UPDATE
                hostname = VALUES(hostname),
                local_ip = VALUES(local_ip),
                last_report = VALUES(last_report)
        ''', (client_id, hostname, local_ip, current_time))

        # 保存进程告警
        alert_data = json.dumps({
            "alerts": alerts,
            "system_summary": system_summary,
            "timestamp": datetime.now(TZ_CST).isoformat()
        }, ensure_ascii=False)

        cursor.execute('''
            INSERT INTO process_alert_records (client_id, hostname, local_ip, alert_data, alert_count)
            VALUES (%s, %s, %s, %s, %s)
        ''', (client_id, hostname, local_ip, alert_data, len(alerts)))

        alert_id = cursor.lastrowid
        conn.commit()

        # 异步 AI 分析（独立线程池，避免 25s+ 的 AI 调用阻塞硬件检测线程）
        ai_executor.submit(
            _async_ai_analyze,
            alert_id, client_id, alert_data
        )

        conn.close()
        return jsonify({'status': 'success', 'message': '进程告警接收成功', 'alert_id': alert_id})

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/process-alerts', methods=['GET'])
@login_required
def get_process_alerts():
    """获取进程告警列表（支持分页、按主机筛选）"""
    try:
        page = int(request.args.get('page', 1))
        per_page = int(request.args.get('per_page', 20))
        client_id = request.args.get('client_id')
        resolved = request.args.get('resolved')

        with get_db_readonly_safe() as conn:
            cursor = conn.cursor()

            query = 'SELECT * FROM process_alert_records WHERE 1=1'
            count_query = 'SELECT COUNT(*) FROM process_alert_records WHERE 1=1'
            params = []

            if client_id:
                query += ' AND client_id = %s'
                count_query += ' AND client_id = %s'
                params.append(client_id)

            if resolved is not None:
                query += ' AND resolved = %s'
                count_query += ' AND resolved = %s'
                params.append(1 if resolved == 'true' else 0)

            # 获取总数
            cursor.execute(count_query, params)
            total = cursor.fetchone()['COUNT(*)']

            # 分页查询
            query += ' ORDER BY created_at DESC LIMIT %s OFFSET %s'
            cursor.execute(query, params + [per_page, (page - 1) * per_page])

            alerts = []
            for row in cursor.fetchall():
                row_dict = dict(row)
                if row_dict.get('alert_data'):
                    row_dict['alert_data'] = json.loads(row_dict['alert_data'])
                if row_dict.get('ai_result'):
                    row_dict['ai_result'] = json.loads(row_dict['ai_result'])
                alerts.append(row_dict)

        return jsonify({
            'status': 'success',
            'data': alerts,
            'pagination': {
                'page': page,
                'per_page': per_page,
                'total': total,
                'pages': (total + per_page - 1) // per_page
            }
        })

    except Exception as e:
        logger.error(f'获取进程告警列表失败: {e}')
        return jsonify({'error': str(e)}), 500


@app.route('/api/process-alerts/<int:alert_id>', methods=['PUT'])
@login_required
def resolve_process_alert(alert_id):
    """标记进程告警为已处理"""
    try:
        with get_db_safe() as conn:
            cursor = conn.cursor()
            cursor.execute('UPDATE process_alert_records SET resolved = 1 WHERE id = %s', (alert_id,))
            if cursor.rowcount == 0:
                return jsonify({'error': '告警记录不存在'}), 404
            conn.commit()
        return jsonify({'status': 'success', 'message': '已标记为已处理'})
    except Exception as e:
        logger.error(f'标记进程告警失败: {e}')
        return jsonify({'error': str(e)}), 500


@app.route('/api/process-alerts/batch-resolve', methods=['PUT'])
@login_required
def batch_resolve_process_alerts():
    """批量标记进程告警为已处理"""
    try:
        data = request.json
        alert_ids = data.get('alert_ids', [])
        if not alert_ids:
            return jsonify({'error': '请选择要处理的告警'}), 400
        with get_db_safe() as conn:
            cursor = conn.cursor()
            placeholders = ','.join(['%s' for _ in alert_ids])
            cursor.execute(f'UPDATE process_alert_records SET resolved = 1 WHERE id IN ({placeholders})', alert_ids)
            affected = cursor.rowcount
            conn.commit()
        return jsonify({'status': 'success', 'affected': affected})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/process-alerts/<int:alert_id>/reanalyze', methods=['POST'])
@login_required
def reanalyze_process_alert(alert_id):
    """手动触发单条告警的 AI 研判"""
    try:
        conn = get_db()
        cursor = conn.cursor()

        cursor.execute('SELECT * FROM process_alert_records WHERE id = %s', (alert_id,))
        alert = cursor.fetchone()
        if not alert:
            conn.close()
            return jsonify({'error': '告警记录不存在'}), 404

        # 获取 AI 配置
        cursor.execute('SELECT * FROM ai_config WHERE id = 1')
        ai_cfg = cursor.fetchone()
        if not ai_cfg or not ai_cfg.get('enabled'):
            conn.close()
            return jsonify({'error': 'AI 未启用'}), 400

        ai_cfg = dict(ai_cfg)
        alert_data = json.loads(alert['alert_data'])

        # 调用 AI 分析
        result = analyze_process_alert(alert_data, ai_cfg)

        if result:
            cursor.execute('''
                UPDATE process_alert_records SET ai_analyzed = 1, ai_result = %s WHERE id = %s
            ''', (json.dumps(result, ensure_ascii=False), alert_id))
            conn.commit()

        conn.close()
        return jsonify({'status': 'success', 'result': result})

    except Exception as e:
        return jsonify({'error': str(e)}), 500


# =================================================================
# AI 研判结果 API（独立查看 / 删除）
# =================================================================

@app.route('/api/ai/results', methods=['GET'])
@login_required
def get_ai_results():
    """获取所有已完成的 AI 研判结果（ai_analyzed=1 且存在 ai_result）"""
    try:
        page = int(request.args.get('page', 1))
        per_page = int(request.args.get('per_page', 20))

        with get_db_readonly_safe() as conn:
            cursor = conn.cursor()

            cursor.execute(
                'SELECT COUNT(*) as total FROM process_alert_records '
                'WHERE ai_analyzed = 1 AND ai_result IS NOT NULL AND ai_result != ""'
            )
            total = cursor.fetchone()['total']

            cursor.execute('''
                SELECT id, client_id, hostname, local_ip, alert_data, alert_count, ai_result, created_at
                FROM process_alert_records
                WHERE ai_analyzed = 1 AND ai_result IS NOT NULL AND ai_result != ''
                ORDER BY created_at DESC
                LIMIT %s OFFSET %s
            ''', (per_page, (page - 1) * per_page))

            results = []
            for row in cursor.fetchall():
                r = dict(row)
                if r.get('alert_data'):
                    r['alert_data'] = json.loads(r['alert_data'])
                if r.get('ai_result'):
                    r['ai_result'] = json.loads(r['ai_result'])
                results.append(r)

        return jsonify({
            'status': 'success',
            'data': results,
            'pagination': {
                'page': page,
                'per_page': per_page,
                'total': total,
                'pages': (total + per_page - 1) // per_page
            }
        })

    except Exception as e:
        logger.error(f'获取 AI 研判结果失败: {e}')
        return jsonify({'error': str(e)}), 500


@app.route('/api/ai/results/<int:alert_id>', methods=['DELETE'])
@login_required
def delete_ai_result(alert_id):
    """删除单条 AI 研判结果（清空 ai_result 并把 ai_analyzed 置 0，保留告警记录可重新分析）"""
    try:
        with get_db_safe() as conn:
            cursor = conn.cursor()
            cursor.execute(
                'UPDATE process_alert_records SET ai_analyzed = 0, ai_result = NULL WHERE id = %s',
                (alert_id,)
            )
            if cursor.rowcount == 0:
                return jsonify({'error': '记录不存在'}), 404
            conn.commit()
        return jsonify({'status': 'success', 'message': '已删除该 AI 研判结果'})
    except Exception as e:
        logger.error(f'删除 AI 研判结果失败: {e}')
        return jsonify({'error': str(e)}), 500


@app.route('/api/ai/results/clear', methods=['POST'])
@login_required
def clear_ai_results():
    """清空所有 AI 研判结果（可选按 client_id 过滤）"""
    try:
        data = request.json or {}
        client_id = data.get('client_id')

        with get_db_safe() as conn:
            cursor = conn.cursor()
            if client_id:
                cursor.execute(
                    'UPDATE process_alert_records SET ai_analyzed = 0, ai_result = NULL '
                    'WHERE client_id = %s AND ai_analyzed = 1',
                    (client_id,)
                )
            else:
                cursor.execute(
                    'UPDATE process_alert_records SET ai_analyzed = 0, ai_result = NULL '
                    'WHERE ai_analyzed = 1'
                )
            affected = cursor.rowcount
            conn.commit()
        return jsonify({'status': 'success', 'message': f'已清空 {affected} 条 AI 研判结果', 'affected': affected})
    except Exception as e:
        logger.error(f'清空 AI 研判结果失败: {e}')
        return jsonify({'error': str(e)}), 500


# =================================================================
# AI 配置 API
# =================================================================

@app.route('/api/ai/config', methods=['GET'])
@login_required
def get_ai_config():
    """获取 AI 配置（api_key 脱敏返回）"""
    try:
        with get_db_readonly_safe() as conn:
            cursor = conn.cursor()
            cursor.execute('SELECT * FROM ai_config WHERE id = 1')
            row = cursor.fetchone()
        if row:
            config = dict(row)
            # 掩码敏感字段
            if config.get('api_key'):
                key = config['api_key']
                if len(key) > 8:
                    config['api_key'] = key[:4] + '****' + key[-4:]
                else:
                    config['api_key'] = '****'
            config['api_key_configured'] = bool(config.get('api_key') and config['api_key'] != '****')
            # Decimal 字段需要转为 float 才能 JSON 序列化
            if 'temperature' in config:
                config['temperature'] = float(config['temperature'])
            return jsonify({'status': 'success', 'data': config})
        return jsonify({'status': 'success', 'data': {
            'enabled': 0, 'api_base_url': 'https://api.openai.com/v1',
            'api_key_configured': False, 'model': 'gpt-4o-mini', 'max_tokens': 2000,
            'temperature': 0.3, 'system_prompt': '', 'auto_analyze': 1
        }})
    except Exception as e:
        logger.error(f'获取AI配置失败: {e}')
        return jsonify({'error': str(e)}), 500


@app.route('/api/ai/config', methods=['PUT'])
@login_required
def update_ai_config():
    """更新 AI 配置"""
    try:
        data = request.json
        with get_db_safe() as conn:
            cursor = conn.cursor()

            # 如果 api_key 是脱敏的，保留原值
            if data.get('api_key') and '****' in data['api_key']:
                cursor.execute('SELECT api_key FROM ai_config WHERE id = 1')
                old = cursor.fetchone()
                data['api_key'] = old['api_key'] if old else ''

            cursor.execute('''
                UPDATE ai_config SET
                    enabled = %s, api_base_url = %s, api_key = %s, model = %s,
                    max_tokens = %s, temperature = %s, system_prompt = %s, auto_analyze = %s,
                    updated_at = CURRENT_TIMESTAMP
                WHERE id = 1
            ''', (
                1 if data.get('enabled') else 0,
                data.get('api_base_url', 'https://api.openai.com/v1'),
                data.get('api_key', ''),
                data.get('model', 'gpt-4o-mini'),
                data.get('max_tokens', 2000),
                float(data.get('temperature', 0.3)),
                data.get('system_prompt', ''),
                1 if data.get('auto_analyze') else 0
            ))
            conn.commit()
        return jsonify({'status': 'success', 'message': 'AI 配置已更新'})
    except Exception as e:
        logger.error(f'更新AI配置失败: {e}')
        return jsonify({'error': str(e)}), 500


@app.route('/api/ai/config/test', methods=['POST'])
@login_required
def test_ai_config():
    """测试 AI 接口连通性"""
    try:
        data = request.json
        success, message = test_ai_connection(data)
        if success:
            return jsonify({'status': 'success', 'message': message})
        else:
            return jsonify({'status': 'error', 'message': message}), 400
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/ai/config/models', methods=['POST'])
@login_required
def fetch_ai_models():
    """获取 AI 模型列表（OpenAI 兼容 /models 端点）"""
    try:
        data = request.json
        api_base_url = data.get('api_base_url', '').rstrip('/')
        api_key = data.get('api_key', '')
        if not api_base_url or not api_key:
            return jsonify({'error': '请填写 API 地址和 Key'}), 400

        url = f"{api_base_url}/models"
        headers = {"Authorization": f"Bearer {api_key}"}
        resp = requests.get(url, headers=headers, timeout=15)
        resp.raise_for_status()
        d = resp.json()
        models = [m.get('id', '') for m in d.get('data', []) if m.get('id')]
        return jsonify({'status': 'success', 'models': sorted(models)})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# =================================================================
# AI 监控主机管理 API
# =================================================================

@app.route('/api/ai/hosts', methods=['GET'])
@login_required
def get_ai_hosts():
    """获取 AI 监控主机列表"""
    try:
        with get_db_readonly_safe() as conn:
            cursor = conn.cursor()
            cursor.execute('SELECT * FROM ai_monitored_hosts ORDER BY added_at DESC')
            hosts = [dict(row) for row in cursor.fetchall()]
        return jsonify({'status': 'success', 'data': hosts})
    except Exception as e:
        logger.error(f'获取AI监控主机失败: {e}')
        return jsonify({'error': str(e)}), 500


@app.route('/api/ai/hosts', methods=['POST'])
@login_required
def add_ai_host():
    """添加 AI 监控主机"""
    try:
        data = request.json
        client_id = data.get('client_id')
        if not client_id:
            return jsonify({'error': '请选择要添加的客户端'}), 400

        hostname = data.get('hostname', '')
        description = data.get('description', '')

        with get_db_safe() as conn:
            cursor = conn.cursor()
            try:
                cursor.execute('''
                    INSERT INTO ai_monitored_hosts (client_id, hostname, description)
                    VALUES (%s, %s, %s)
                ''', (client_id, hostname, description))
                conn.commit()
                host_id = cursor.lastrowid
                return jsonify({'status': 'success', 'host_id': host_id})
            except pymysql.err.IntegrityError:
                return jsonify({'error': '该主机已在监控列表中'}), 400
    except Exception as e:
        logger.error(f'添加AI监控主机失败: {e}')
        return jsonify({'error': str(e)}), 500


@app.route('/api/ai/hosts/<int:host_id>', methods=['DELETE'])
@login_required
def delete_ai_host(host_id):
    """移除 AI 监控主机"""
    try:
        with get_db_safe() as conn:
            cursor = conn.cursor()
            cursor.execute('DELETE FROM ai_monitored_hosts WHERE id = %s', (host_id,))
            if cursor.rowcount == 0:
                return jsonify({'error': '主机不存在'}), 404
        conn.commit()
        conn.close()
        return jsonify({'status': 'success', 'message': '已移除'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# =================================================================
# 主机探测模块 API
# =================================================================

@app.route('/api/probe/targets', methods=['GET'])
@login_required
def get_probe_targets():
    """获取探测目标列表（支持分组筛选）"""
    try:
        group_id = request.args.get('group_id')
        with get_db_readonly_safe() as conn:
            cursor = conn.cursor()
            if group_id:
                cursor.execute('SELECT * FROM probe_targets WHERE group_id = %s ORDER BY created_at DESC', (group_id,))
            else:
                cursor.execute('SELECT * FROM probe_targets ORDER BY created_at DESC')
            targets = [dict(row) for row in cursor.fetchall()]
        return jsonify({'status': 'success', 'data': targets})
    except Exception as e:
        logger.error(f'获取探测目标失败: {e}')
        return jsonify({'error': str(e)}), 500


@app.route('/api/probe/targets', methods=['POST'])
@login_required
def add_probe_target():
    """添加探测目标"""
    try:
        data = request.json
        name = data.get('name', '').strip()
        host = data.get('host', '').strip()
        port = int(data.get('port', 80))
        protocol = data.get('protocol', 'http')
        path = data.get('path', '/')
        description = data.get('description', '')
        group_id = data.get('group_id')

        if not name or not host:
            return jsonify({'error': '名称和地址不能为空'}), 400

        # SSRF 防护
        if not validate_probe_host(host):
            return jsonify({'error': '该地址不允许被探测（安全限制）'}), 400

        with get_db_safe() as conn:
            cursor = conn.cursor()
            cursor.execute('''
                INSERT INTO probe_targets (name, host, port, protocol, path, description, group_id)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
            ''', (name, host, port, protocol, path, description, group_id))
            conn.commit()
            target_id = cursor.lastrowid
        return jsonify({'status': 'success', 'target_id': target_id})
    except Exception as e:
        logger.error(f'添加探测目标失败: {e}')
        return jsonify({'error': str(e)}), 500


@app.route('/api/probe/targets/<int:target_id>', methods=['PUT'])
@login_required
def update_probe_target(target_id):
    try:
        data = request.json
        with get_db_safe() as conn:
            cursor = conn.cursor()
            cursor.execute('''UPDATE probe_targets SET name=%s, host=%s, port=%s, protocol=%s, path=%s, description=%s, enabled=%s, group_id=%s WHERE id=%s''',
                (data.get('name'), data.get('host'), int(data.get('port',80)), data.get('protocol','http'), data.get('path','/'), data.get('description',''), 1 if data.get('enabled',True) else 0, data.get('group_id'), target_id))
            conn.commit()
        return jsonify({'status': 'success'})
    except Exception as e:
        logger.error(f'更新探测目标失败: {e}')
        return jsonify({'error': str(e)}), 500


@app.route('/api/probe/targets/<int:target_id>', methods=['DELETE'])
@login_required
def delete_probe_target(target_id):
    try:
        with get_db_safe() as conn:
            cursor = conn.cursor()
            cursor.execute('DELETE FROM probe_targets WHERE id = %s', (target_id,))
            conn.commit()
        return jsonify({'status': 'success'})
    except Exception as e:
        logger.error(f'删除探测目标失败: {e}')
        return jsonify({'error': str(e)}), 500


@app.route('/api/probe/targets/<int:target_id>/group', methods=['PUT'])
@login_required
def assign_probe_group(target_id):
    """分配探测目标到分组"""
    try:
        data = request.json
        group_id = data.get('group_id')
        with get_db_safe() as conn:
            cursor = conn.cursor()
            cursor.execute('UPDATE probe_targets SET group_id = %s WHERE id = %s', (group_id, target_id))
            conn.commit()
        return jsonify({'status': 'success'})
    except Exception as e:
        logger.error(f'分配探测目标分组失败: {e}')
        return jsonify({'error': str(e)}), 500


@app.route('/api/probe/targets/<int:target_id>/test', methods=['POST'])
@login_required
def test_probe_target(target_id):
    try:
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute('SELECT * FROM probe_targets WHERE id = %s', (target_id,))
        target = cursor.fetchone()
        conn.close()
        if not target:
            return jsonify({'error': '目标不存在'}), 404
        result = _probe_single_target(dict(target))
        return jsonify({'status': 'success', 'result': result})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/probe/alerts', methods=['GET'])
@login_required
def get_probe_alerts():
    try:
        page = int(request.args.get('page', 1))
        per_page = int(request.args.get('per_page', 50))
        resolved = request.args.get('resolved')
        with get_db_readonly_safe() as conn:
            cursor = conn.cursor()
            query = 'SELECT * FROM probe_alerts WHERE 1=1'
            count_q = 'SELECT COUNT(*) as total FROM probe_alerts WHERE 1=1'
            params = []
            if resolved is not None:
                query += ' AND resolved = %s'
                count_q += ' AND resolved = %s'
                params.append(1 if resolved == 'true' else 0)
            cursor.execute(count_q, params)
            total = cursor.fetchone()['total']
            query += ' ORDER BY created_at DESC LIMIT %s OFFSET %s'
            cursor.execute(query, params + [per_page, (page - 1) * per_page])
            alerts = [dict(row) for row in cursor.fetchall()]
        return jsonify({'status': 'success', 'data': alerts, 'pagination': {'page': page, 'per_page': per_page, 'total': total, 'pages': (total + per_page - 1) // per_page}})
    except Exception as e:
        logger.error(f'获取探测告警失败: {e}')
        return jsonify({'error': str(e)}), 500


@app.route('/api/probe/alerts/<int:alert_id>/resolve', methods=['PUT'])
@login_required
def resolve_probe_alert(alert_id):
    try:
        with get_db_safe() as conn:
            cursor = conn.cursor()
            cursor.execute('UPDATE probe_alerts SET resolved = 1 WHERE id = %s', (alert_id,))
            conn.commit()
        return jsonify({'status': 'success'})
    except Exception as e:
        logger.error(f'标记探测告警失败: {e}')
        return jsonify({'error': str(e)}), 500


@app.route('/api/probe/alerts/<int:alert_id>', methods=['DELETE'])
@login_required
def delete_probe_alert(alert_id):
    try:
        with get_db_safe() as conn:
            cursor = conn.cursor()
            cursor.execute('DELETE FROM probe_alerts WHERE id = %s', (alert_id,))
            conn.commit()
        return jsonify({'status': 'success'})
    except Exception as e:
        logger.error(f'删除探测告警失败: {e}')
        return jsonify({'error': str(e)}), 500


@app.route('/api/probe/alerts/batch-resolve', methods=['PUT'])
@login_required
def batch_resolve_probe_alerts():
    try:
        data = request.json
        ids = data.get('ids', [])
        if not ids:
            return jsonify({'error': '请选择告警'}), 400
        with get_db_safe() as conn:
            cursor = conn.cursor()
            ph = ','.join(['%s' for _ in ids])
            cursor.execute(f'UPDATE probe_alerts SET resolved = 1 WHERE id IN ({ph})', ids)
            affected = cursor.rowcount
            conn.commit()
        return jsonify({'status': 'success', 'affected': affected})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


def _probe_single_target(target):
    import socket as _socket
    import subprocess as _sp
    import platform as _pf
    host = target['host']
    port = int(target.get('port', 80))
    protocol = target.get('protocol', 'http')
    path = target.get('path', '/')
    result = {'host': host, 'port': port, 'protocol': protocol, 'status': 'unknown', 'status_code': None, 'response_time': None, 'error': None}
    start = time.time()
    try:
        if protocol == 'ping':
            # ICMP Ping
            param = '-n' if _pf.system().lower() == 'windows' else '-c'
            r = _sp.run(['ping', param, '1', '-w' if _pf.system().lower() == 'windows' else '-W', '5', host],
                        capture_output=True, text=True, timeout=10)
            elapsed = int((time.time() - start) * 1000)
            result['response_time'] = elapsed
            result['status'] = 'online' if r.returncode == 0 else 'offline'
            if r.returncode != 0:
                result['error'] = 'Ping 失败'
        elif protocol in ('http', 'https'):
            resp = requests.get(f"{protocol}://{host}:{port}{path}", timeout=10, verify=False)
            result['status_code'] = resp.status_code
            result['response_time'] = int((time.time() - start) * 1000)
            result['status'] = 'online' if resp.status_code < 500 else 'offline'
        else:
            sock = _socket.socket(_socket.AF_INET, _socket.SOCK_STREAM)
            sock.settimeout(10)
            r = sock.connect_ex((host, port))
            result['response_time'] = int((time.time() - start) * 1000)
            sock.close()
            result['status'] = 'online' if r == 0 else 'offline'
    except Exception as e:
        result['status'] = 'offline'
        result['error'] = str(e)[:200]
    return result


def probe_all_targets():
    try:
        with get_db_safe() as conn:
            cursor = conn.cursor()
            cursor.execute('SELECT * FROM probe_targets WHERE enabled = 1')
            targets = [dict(row) for row in cursor.fetchall()]
            if not targets:
                return
            for target in targets:
                result = _probe_single_target(target)
                cursor.execute('UPDATE probe_targets SET last_status=%s, last_probe_time=NOW(), last_response_time=%s WHERE id=%s',
                    (result['status'], result.get('response_time'), target['id']))
                if result['status'] == 'offline':
                    cursor.execute('SELECT id FROM probe_alerts WHERE target_id = %s AND resolved = 0 LIMIT 1', (target['id'],))
                    if not cursor.fetchone():
                        cursor.execute('INSERT INTO probe_alerts (target_id, target_name, target_host, alert_type, error_message) VALUES (%s,%s,%s,%s,%s)',
                            (target['id'], target['name'], target['host'], 'offline', result.get('error','')))
                else:
                    cursor.execute('UPDATE probe_alerts SET resolved = 1 WHERE target_id = %s AND resolved = 0 AND alert_type = "offline"', (target['id'],))
            conn.commit()
    except Exception as e:
        logger.warning(f'探测目标失败: {e}')


# =================================================================
# 后台定时任务
# =================================================================

def cleanup_old_records():
    import time as tmod
    while True:
        try:
            with get_db_safe() as conn:
                cursor = conn.cursor()

                # 从 clients 表拿 client_id（约 200 行，秒级），避免 GROUP BY 全索引扫描
                # （20 万行 LONGTEXT 大表上 GROUP BY 全索引扫描实测 >224s，点查点删则秒级）
                cursor.execute('SELECT client_id FROM clients')
                client_ids = [r['client_id'] for r in cursor.fetchall()]

                # 清理硬件历史表：每客户端保留最近 10 条（点查 + 点删，走 idx_client_id 索引）
                for cid in client_ids:
                    try:
                        cursor.execute('SELECT id FROM hardware_history WHERE client_id=%s ORDER BY id DESC LIMIT 9,1', (cid,))
                        row = cursor.fetchone()
                        if row:
                            cursor.execute('DELETE FROM hardware_history WHERE client_id=%s AND id < %s', (cid, row['id']))
                    except Exception:
                        pass

                # 清理上报明细表：每客户端保留最近 10 条
                # （全项目对 hardware_reports 只查最新一条，历史由 hardware_history 承担，故仅保留少量冗余兜底）
                for cid in client_ids:
                    try:
                        cursor.execute('SELECT id FROM hardware_reports WHERE client_id=%s ORDER BY id DESC LIMIT 9,1', (cid,))
                        row = cursor.fetchone()
                        if row:
                            cursor.execute('DELETE FROM hardware_reports WHERE client_id=%s AND id < %s', (cid, row['id']))
                    except Exception:
                        pass

                conn.commit()
        except Exception as e:
            logger.warning(f'清理旧记录失败: {e}')
        tmod.sleep(1800)


def create_app():
    """初始化应用（供 gunicorn 使用）"""
    init_db_pool()
    init_tables()
    init_redis()

    try:
        from celery_app import init_celery
        init_celery()
    except Exception as e:
        logger.warning(f'Celery 初始化失败: {e}')

    import threading
    threading.Thread(target=cleanup_old_records, daemon=True).start()
    logger.info('cleanup started')

    def _probe_loop():
        while True:
            try:
                probe_all_targets()
            except Exception:
                pass
            time.sleep(30)
    threading.Thread(target=_probe_loop, daemon=True).start()
    logger.info('probe started (30s)')

    return app


if __name__ == '__main__':
    create_app()
    logger.info("=" * 50)
    logger.info("HwMon Server v5.0.1 - http://localhost:5000")
    logger.info("=" * 50)
    try:
        from waitress import serve
        serve(app, host='0.0.0.0', port=5000, threads=50, connection_limit=3000)
    except ImportError:
        app.run(host='0.0.0.0', port=5000, debug=True)


